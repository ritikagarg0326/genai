"""
main.py  —  🎉 GenAI Ads Reddit Scraper  (Fun Edition!)
══════════════════════════════════════════════════════
Run:  python main.py
"""

from __future__ import annotations

import sys, os, threading, queue, time, random, re, math
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.scrolledtext import ScrolledText

try:
    import requests
except ImportError:
    os.system(f"{sys.executable} -m pip install requests -q"); import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
#  CANDY PALETTE
# ══════════════════════════════════════════════════════════════════════════════
C = {
    "sky":      "#eef6ff",
    "cloud":    "#ffffff",
    "lavender": "#f3eeff",
    "mint":     "#e5fbf4",
    "peach":    "#fff3ec",
    "lemon":    "#fffce8",
    "pink":     "#ff6eb4",
    "purple":   "#9b59ff",
    "blue":     "#3da9fc",
    "teal":     "#00c9a7",
    "orange":   "#ff8c42",
    "yellow":   "#ffd93d",
    "red":      "#ff4757",
    "green":    "#2ecc71",
    "text":     "#2d2d4e",
    "muted":    "#8888aa",
    "white":    "#ffffff",
    "shadow":   "#d8d8f0",
}
RAINBOW = [C["pink"], C["orange"], C["yellow"], C["green"], C["blue"], C["purple"]]

DEFAULT_SUBREDDITS = [
    "advertising","SocialMediaMarketing","FacebookAds","PPC",
    "digital_marketing","GoogleAds","marketing","ArtificialIntelligence",
]
DEFAULT_KEYWORDS = [
    "generative AI advertising","AI ads","GenAI marketing",
    "AI creative ads","LLM advertising","AI ad generation",
    "ChatGPT ads","AI content marketing",
]
SORT_OPTIONS = ["relevance","new","top","comments"]
TIME_OPTIONS = ["all","year","month","week","day"]
COMMENT_COLOURS = ["FFF9C4","FFE0B2","F8BBD0","E1BEE7","C8E6C9","B3E5FC","DCEDC8","F0F4C3","FFE0B2","FCE4EC"]


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPER
# ══════════════════════════════════════════════════════════════════════════════
SEARCH_URL = "https://www.reddit.com/r/{sub}/search.json"

def make_session():
    s = requests.Session()
    s.headers.update({"User-Agent":"Mozilla/5.0 (compatible; GenAI-Ads-Scraper/1.0; research)","Accept":"application/json"})
    return s

def search_subreddit(session, sub, query, sort, time_filter, limit, log_q):
    params={"q":query,"sort":sort,"restrict_sr":"1","t":time_filter,"limit":min(limit,100),"raw_json":"1"}
    try:
        r=session.get(SEARCH_URL.format(sub=sub),params=params,timeout=20)
        if r.status_code==429:
            wait=int(r.headers.get("Retry-After",30))
            log_q.put(("warn",f"⏳ Rate limited — waiting {wait}s…"))
            time.sleep(wait); return []
        if r.status_code!=200:
            log_q.put(("warn",f"😬 r/{sub} HTTP {r.status_code}")); return []
        return r.json().get("data",{}).get("children",[])
    except Exception as e:
        log_q.put(("warn",f"😬 r/{sub} error: {e}")); return []

def fetch_top_comments(session, permalink, n, log_q):
    url=f"https://www.reddit.com{permalink}.json?limit={n+5}&sort=top&raw_json=1"
    try:
        r=session.get(url,timeout=20)
        if r.status_code!=200: return []
        children=r.json()[1]["data"]["children"]
        comments=[]
        for child in children:
            body=child.get("data",{}).get("body","")
            if child.get("kind")=="t1" and body and body not in("[deleted]","[removed]"):
                comments.append(re.sub(r"\s+"," ",body).strip()[:500])
            if len(comments)>=n: break
        return comments
    except: return []

def scrape(config, log_q, progress_q):
    cutoff=datetime.now(timezone.utc).replace(tzinfo=None)-timedelta(days=config["days_back"])
    records,seen_ids,session=[],set(),make_session()
    subs=config["subreddits"]
    for sub_i,sub in enumerate(subs,1):
        log_q.put(("head",f"🚀 Blasting into r/{sub}  ({sub_i}/{len(subs)})…"))
        sub_count=0
        for keyword in config["keywords"]:
            if sub_count>=config["max_posts"]: break
            children=search_subreddit(session,sub,keyword,config["sort_by"],config["time_filter"],25,log_q)
            time.sleep(random.uniform(1.2,2.5))
            for child in children:
                if sub_count>=config["max_posts"]: break
                post=child.get("data",{})
                pid=post.get("id","")
                if not pid or pid in seen_ids: continue
                seen_ids.add(pid)
                ts=post.get("created_utc",0)
                dt=datetime.utcfromtimestamp(ts) if ts else None
                if dt and dt<cutoff: continue
                title=post.get("title","").strip()
                permalink=post.get("permalink","")
                score=int(post.get("score",0))
                n_comments=int(post.get("num_comments",0))
                log_q.put(("ok",f"  ✨ {title[:55]}…  👍{score}  💬{n_comments}"))
                comments=[]
                if permalink and n_comments>0:
                    comments=fetch_top_comments(session,permalink,config["max_comments"],log_q)
                    time.sleep(random.uniform(1.0,2.0))
                record={"subreddit":f"r/{sub}","post_title":title,
                        "post_url":f"https://www.reddit.com{permalink}",
                        "score":score,"date":dt.strftime("%Y-%m-%d") if dt else ""}
                for i in range(1,config["max_comments"]+1):
                    record[f"comment_{i}"]=comments[i-1] if i<=len(comments) else ""
                records.append(record); sub_count+=1
        log_q.put(("ok",f"  🎯 r/{sub}: {sub_count} posts collected!"))
        progress_q.put(sub_i/len(subs))
    log_q.put(("star",f"🎉 WOW! Total posts scraped: {len(records)}"))
    return records

def export_excel(records, filepath, max_comments):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="GenAI Ads — Reddit"
    HF=PatternFill("solid",fgColor="1F4E79"); AF=PatternFill("solid",fgColor="F2F7FC")
    HN=Font(bold=True,color="FFFFFF",size=11); LN=Font(color="0563C1",underline="single",size=10)
    NF=Font(size=10)
    CE=Alignment(horizontal="center",vertical="top",wrap_text=False)
    LW=Alignment(horizontal="left",vertical="top",wrap_text=True)
    TN=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    cols=["Subreddit","Post Title","Post Link","Score","Date"]+[f"Comment {i}" for i in range(1,max_comments+1)]
    ws.append(cols)
    for ci in range(1,len(cols)+1):
        c=ws.cell(row=1,column=ci); c.fill=HF; c.font=HN; c.alignment=CE; c.border=TN
    ws.row_dimensions[1].height=28
    for ri,rec in enumerate(records,start=2):
        rf=AF if ri%2==0 else PatternFill("solid",fgColor="FFFFFF")
        def cell(col,val,fill=None,font=None,align=None):
            c=ws.cell(row=ri,column=col,value=val)
            c.fill=fill or rf; c.font=font or NF; c.alignment=align or CE; c.border=TN
            return c
        cell(1,rec["subreddit"],fill=PatternFill("solid",fgColor="D6E4F0"),font=Font(bold=True,size=10))
        cell(2,rec["post_title"],align=LW)
        c=cell(3,"🔗 Open Post",font=LN); c.hyperlink=rec["post_url"]
        cell(4,rec["score"])
        cell(5,rec["date"])
        for i in range(1,max_comments+1):
            txt=rec.get(f"comment_{i}",""); col=COMMENT_COLOURS[(i-1)%len(COMMENT_COLOURS)]
            cell(5+i,txt,fill=PatternFill("solid",fgColor=col) if txt else rf,align=LW)
        ws.row_dimensions[ri].height=60
    widths=[16,45,14,8,12]+[35]*max_comments
    for ci,w in enumerate(widths,start=1): ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions; wb.save(filepath)


# ══════════════════════════════════════════════════════════════════════════════
#  FLOATING SPARKLES CANVAS
# ══════════════════════════════════════════════════════════════════════════════
class SparkleCanvas(tk.Canvas):
    SYMBOLS=["★","✦","✧","◆","✿","❋","♦","●"]
    def __init__(self,parent,**kw):
        super().__init__(parent,highlightthickness=0,**kw)
        self._stars=[]
        self.after(200,self._init)
    def _init(self):
        w,h=max(self.winfo_width(),800),max(self.winfo_height(),100)
        for _ in range(28):
            self._stars.append({
                "x":random.uniform(0,w),"y":random.uniform(0,h),
                "vy":random.uniform(-0.5,-0.15),"vx":random.uniform(-0.2,0.2),
                "size":random.randint(9,20),"sym":random.choice(self.SYMBOLS),
                "color":random.choice(RAINBOW),"phase":random.uniform(0,6.28),
                "dphase":random.uniform(0.03,0.09),
            })
        self._tick()
    def _tick(self):
        self.delete("sp")
        w,h=max(self.winfo_width(),800),max(self.winfo_height(),100)
        for s in self._stars:
            s["x"]=(s["x"]+s["vx"])%w; s["y"]+=s["vy"]
            if s["y"]<-24: s["y"]=h+10
            s["phase"]+=s["dphase"]
            sz=s["size"]+int(3*math.sin(s["phase"]))
            self.create_text(s["x"],s["y"],text=s["sym"],font=("Arial",sz),fill=s["color"],tags="sp")
        self.after(40,self._tick)


# ══════════════════════════════════════════════════════════════════════════════
#  PILL BUTTON (candy-style, pulsing shadow)
# ══════════════════════════════════════════════════════════════════════════════
class PillButton(tk.Canvas):
    def __init__(self,parent,text,command,color,**kw):
        super().__init__(parent,highlightthickness=0,cursor="hand2",bg=parent.cget("bg"),**kw)
        self._text=text; self._cmd=command; self._color=color
        self._phase=0.0; self._pressed=False
        self.bind("<Button-1>",self._press)
        self.bind("<ButtonRelease-1>",self._release)
        self.bind("<Configure>",lambda _:self._draw())
        self.after(20,self._animate)

    def _lgt(self,h,a=55):
        h=h.lstrip("#"); r,g,b=int(h[0:2],16),int(h[2:4],16),int(h[4:6],16)
        return f"#{min(255,r+a):02x}{min(255,g+a):02x}{min(255,b+a):02x}"
    def _drk(self,h,a=30):
        h=h.lstrip("#"); r,g,b=int(h[0:2],16),int(h[2:4],16),int(h[4:6],16)
        return f"#{max(0,r-a):02x}{max(0,g-a):02x}{max(0,b-a):02x}"

    def _pill(self,x1,y1,x2,y2,r,**kw):
        self.create_arc(x1,y1,x1+2*r,y1+2*r,start=90,extent=90,style="pieslice",**kw)
        self.create_arc(x2-2*r,y1,x2,y1+2*r,start=0,extent=90,style="pieslice",**kw)
        self.create_arc(x1,y2-2*r,x1+2*r,y2,start=180,extent=90,style="pieslice",**kw)
        self.create_arc(x2-2*r,y2-2*r,x2,y2,start=270,extent=90,style="pieslice",**kw)
        self.create_rectangle(x1+r,y1,x2-r,y2,**kw)
        self.create_rectangle(x1,y1+r,x2,y2-r,**kw)

    def _draw(self):
        self.delete("all")
        w,h=self.winfo_width(),self.winfo_height()
        if w<10 or h<10: return
        g=int(5+4*math.sin(self._phase)); r=h//2
        fill=self._drk(self._color) if self._pressed else self._color
        self._pill(g,g,w-2,h-2,r,fill="#c0bce8",outline="")
        self._pill(0,0,w-g-2,h-g-2,r,fill=fill,outline="")
        self._pill(8,5,w-g-10,h//2-2,r-4,fill=self._lgt(fill,70),outline="")
        self.create_text((w-g)//2,(h-g)//2,text=self._text,
                         font=("Comic Sans MS",14,"bold"),fill="#ffffff")

    def _animate(self):
        self._phase+=0.06; self._draw(); self.after(30,self._animate)

    def _press(self,_): self._pressed=True; self._cmd(); self._draw()
    def _release(self,_): self._pressed=False; self._draw()

    def update_label(self,t): self._text=t; self._draw()
    def update_color(self,c): self._color=c; self._draw()


# ══════════════════════════════════════════════════════════════════════════════
#  TOOLTIP
# ══════════════════════════════════════════════════════════════════════════════
def attach_tip(widget,text):
    tip=[None]
    def show(_):
        if tip[0]: return
        w=tk.Toplevel(widget); w.wm_overrideredirect(True)
        x=widget.winfo_rootx()+22; y=widget.winfo_rooty()+widget.winfo_height()+4
        w.wm_geometry(f"+{x}+{y}")
        tk.Label(w,text=text,bg="#fffde7",fg=C["text"],font=("Arial",9),
                 relief="solid",bd=1,padx=10,pady=6,wraplength=300).pack()
        tip[0]=w
    def hide(_):
        if tip[0]: tip[0].destroy(); tip[0]=None
    widget.bind("<Enter>",show); widget.bind("<Leave>",hide)


# ══════════════════════════════════════════════════════════════════════════════
#  CARD HELPER
# ══════════════════════════════════════════════════════════════════════════════
def make_card(parent,bg,title_emoji,title,subtitle,pady=10):
    outer=tk.Frame(parent,bg=bg,relief="flat")
    outer.pack(fill="x",pady=(0,pady))
    inner=tk.Frame(outer,bg=bg)
    inner.pack(fill="x",padx=14,pady=12)
    hdr=tk.Frame(inner,bg=bg); hdr.pack(fill="x",pady=(0,4))
    tk.Label(hdr,text=title_emoji,font=("Arial",18),bg=bg).pack(side="left",padx=(0,6))
    tk.Label(hdr,text=title,font=("Comic Sans MS",12,"bold"),bg=bg,fg=C["text"]).pack(side="left")
    if subtitle:
        tk.Label(hdr,text=f"  — {subtitle}",font=("Arial",8),bg=bg,fg=C["muted"]).pack(side="left")
    return inner


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN APP
# ══════════════════════════════════════════════════════════════════════════════
class FunScraperApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("🎉 GenAI Ads Reddit Scraper — FUN EDITION!")
        self.geometry("1090x880"); self.minsize(920,740)
        self.configure(bg=C["sky"])
        self.update_idletasks()
        sw,sh=self.winfo_screenwidth(),self.winfo_screenheight()
        self.geometry(f"1090x880+{(sw-1090)//2}+{(sh-880)//2}")
        self._log_q=queue.Queue(); self._prog_q=queue.Queue(); self._running=False
        self._build(); self._poll()

    # ── BUILD ─────────────────────────────────────────────────────────────────
    def _build(self):
        self._hdr(); self._body(); self._logpane()

    def _hdr(self):
        hdr=tk.Frame(self,bg=C["sky"],height=96); hdr.pack(fill="x"); hdr.pack_propagate(False)
        sc=SparkleCanvas(hdr,bg=C["sky"]); sc.place(x=0,y=0,relwidth=1,relheight=1)
        tf=tk.Frame(hdr,bg=C["sky"]); tf.place(relx=0.5,rely=0.5,anchor="center")
        row=tk.Frame(tf,bg=C["sky"]); row.pack()
        for txt,col in [("🌟","#ffd93d"),(" GenAI ",C["purple"]),("Ads ",C["pink"]),
                        ("Reddit ",C["blue"]),("Scraper ","#ff8c42"),("🚀",C["green"])]:
            tk.Label(row,text=txt,font=("Comic Sans MS",26,"bold"),bg=C["sky"],fg=col).pack(side="left")
        tk.Label(tf,text="✨  Fill in the form below, then click  GENERATE REPORT  to collect Reddit posts! ✨",
                 font=("Comic Sans MS",10),bg=C["sky"],fg=C["muted"]).pack()

    def _body(self):
        outer=tk.Frame(self,bg=C["sky"]); outer.pack(fill="both",expand=True,padx=16,pady=(6,0))
        cv=tk.Canvas(outer,bg=C["sky"],highlightthickness=0)
        vsb=ttk.Scrollbar(outer,orient="vertical",command=cv.yview)
        cv.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right",fill="y"); cv.pack(side="left",fill="both",expand=True)
        sf=tk.Frame(cv,bg=C["sky"]); fid=cv.create_window((0,0),window=sf,anchor="nw")
        sf.bind("<Configure>",lambda _:cv.configure(scrollregion=cv.bbox("all")))
        cv.bind("<Configure>",lambda e:cv.itemconfig(fid,width=e.width))
        cv.bind_all("<MouseWheel>",lambda e:cv.yview_scroll(int(-1*(e.delta/120)),"units"))
        self._form(sf)

    def _form(self,parent):
        cols=tk.Frame(parent,bg=C["sky"]); cols.pack(fill="x",pady=(4,0))
        left=tk.Frame(cols,bg=C["sky"]); right=tk.Frame(cols,bg=C["sky"])
        left.pack(side="left",fill="both",expand=True,padx=(0,8))
        right.pack(side="left",fill="both",expand=True)

        # ── LEFT: Subreddits ──────────────────────────────────────────────────
        si=make_card(left,C["lavender"],"📋","Subreddits","comma-separated, no r/ needed")
        self._note(si,"Type subreddit names separated by commas.  Example:  advertising, FacebookAds, GoogleAds",C["lavender"])
        self.sub_text=self._textbox(si,", ".join(DEFAULT_SUBREDDITS),C["lavender"],4)

        # ── LEFT: Keywords ────────────────────────────────────────────────────
        ki=make_card(left,C["mint"],"🔍","Search Keywords","one keyword or phrase per line")
        self._note(ki,"Each keyword is searched in every subreddit above. More keywords = more posts found.",C["mint"])
        self.kw_text=self._textbox(ki,"\n".join(DEFAULT_KEYWORDS),C["mint"],7)

        # ── RIGHT: Parameters ─────────────────────────────────────────────────
        pi=make_card(right,C["peach"],"⚙️","Scrape Settings","control how much data to collect")
        g=tk.Frame(pi,bg=C["peach"]); g.pack(fill="x")
        g.columnconfigure(1,weight=1)

        self.max_posts_var=tk.IntVar(value=50)
        self.max_comments_var=tk.IntVar(value=10)
        self.days_back_var=tk.IntVar(value=180)
        self.sort_var=tk.StringVar(value="relevance")
        self.time_var=tk.StringVar(value="all")

        spinfields=[
            ("📦 Max Posts / Subreddit",self.max_posts_var,1,500,
             "How many posts per subreddit?\n50 is a great start! More posts = more time."),
            ("💬 Comments per Post",self.max_comments_var,1,25,
             "Each comment gets its OWN column in Excel.\n10 = 10 comment columns."),
            ("📅 Days Back",self.days_back_var,0,3650,
             "Only grab posts from last N days.\n30=last month  180=6 months  365=1 year  0=no filter"),
        ]
        for rn,(lbl,var,lo,hi,tip) in enumerate(spinfields):
            l=tk.Label(g,text=lbl,font=("Comic Sans MS",10,"bold"),bg=C["peach"],fg=C["text"],anchor="w")
            l.grid(row=rn*2,column=0,sticky="w",pady=(10,0)); attach_tip(l,tip)
            sb=tk.Spinbox(g,textvariable=var,from_=lo,to=hi,width=8,
                          bg=C["white"],fg=C["text"],relief="flat",
                          font=("Comic Sans MS",11,"bold"),
                          buttonbackground=C["orange"],
                          highlightthickness=2,highlightcolor=C["orange"],
                          insertbackground=C["text"])
            sb.grid(row=rn*2,column=1,sticky="w",pady=(10,0),padx=(14,0))
            tk.Label(g,text=tip.split("\n")[0],font=("Arial",8),
                     bg=C["peach"],fg=C["muted"],anchor="w").grid(row=rn*2+1,column=0,columnspan=2,sticky="w",padx=4)

        dropfields=[
            ("🔀 Sort Results By",self.sort_var,SORT_OPTIONS,
             "relevance = best keyword match\nnew = newest posts first\ntop = highest upvotes\ncomments = most discussed"),
            ("🕐 Reddit Time Filter",self.time_var,TIME_OPTIONS,
             "Reddit's own time filter (server-side).\nUse 'all' combined with Days Back for best control."),
        ]
        for rn,(lbl,var,opts,tip) in enumerate(dropfields,start=len(spinfields)):
            l=tk.Label(g,text=lbl,font=("Comic Sans MS",10,"bold"),bg=C["peach"],fg=C["text"],anchor="w")
            l.grid(row=rn*2,column=0,sticky="w",pady=(10,0)); attach_tip(l,tip)
            dd=ttk.Combobox(g,textvariable=var,values=opts,state="readonly",
                            width=14,font=("Comic Sans MS",10))
            dd.grid(row=rn*2,column=1,sticky="w",pady=(10,0),padx=(14,0))
            tk.Label(g,text=tip.split("\n")[0],font=("Arial",8),
                     bg=C["peach"],fg=C["muted"],anchor="w").grid(row=rn*2+1,column=0,columnspan=2,sticky="w",padx=4)

        # ── RIGHT: Output ─────────────────────────────────────────────────────
        oi=make_card(right,C["lemon"],"💾","Output File","where should the Excel be saved?")
        self._note(oi,"📁 Folder to save in  (click Browse or type a path)",C["lemon"])
        pr=tk.Frame(oi,bg=C["lemon"]); pr.pack(fill="x",pady=(4,0))
        self.out_dir_var=tk.StringVar(value=str(Path.cwd()/"data"))
        pe=tk.Entry(pr,textvariable=self.out_dir_var,bg=C["white"],fg=C["text"],
                    relief="flat",font=("Comic Sans MS",10),bd=0,
                    highlightthickness=2,highlightcolor=C["yellow"],insertbackground=C["text"])
        pe.pack(side="left",fill="x",expand=True,ipady=6,ipadx=6)
        attach_tip(pe,"The folder where the Excel file will be saved.\nIt will be created if it doesn't exist yet.")
        tk.Button(pr,text="📂 Browse…",command=self._browse,
                  bg=C["orange"],fg=C["white"],relief="flat",
                  font=("Comic Sans MS",10,"bold"),
                  activebackground=C["red"],activeforeground=C["white"],
                  padx=8,pady=4,cursor="hand2",bd=0).pack(side="left",padx=(6,0))
        self._note(oi,"📄 File name  (must end with .xlsx)",C["lemon"])
        self.fname_var=tk.StringVar(value="genai_ads_reddit.xlsx")
        fe=tk.Entry(oi,textvariable=self.fname_var,bg=C["white"],fg=C["text"],
                    relief="flat",font=("Comic Sans MS",10),bd=0,
                    highlightthickness=2,highlightcolor=C["yellow"],insertbackground=C["text"])
        fe.pack(fill="x",ipady=6,ipadx=6,pady=(4,0))
        attach_tip(fe,"Name for your Excel report file.\nMake sure it ends with .xlsx")

        # ── BIG CANDY BUTTON ──────────────────────────────────────────────────
        bf=tk.Frame(right,bg=C["sky"]); bf.pack(fill="x",pady=(8,0))
        self.run_btn=PillButton(bf,"🚀  GENERATE REPORT!",self._start,C["purple"],height=60)
        self.run_btn.pack(fill="x")

        # progress
        pf=tk.Frame(right,bg=C["sky"]); pf.pack(fill="x",pady=(8,0))
        self.prog_var=tk.DoubleVar(value=0)
        sty=ttk.Style(); sty.theme_use("clam")
        sty.configure("Candy.Horizontal.TProgressbar",
                      troughcolor=C["shadow"],background=C["pink"],
                      lightcolor=C["pink"],darkcolor=C["pink"],
                      bordercolor=C["sky"],thickness=16)
        ttk.Progressbar(pf,variable=self.prog_var,maximum=1.0,
                        style="Candy.Horizontal.TProgressbar").pack(fill="x")
        self.status_lbl=tk.Label(pf,text="👋 Ready! Fill in the form and hit the big button!",
                                 bg=C["sky"],fg=C["muted"],
                                 font=("Comic Sans MS",9),wraplength=500,justify="left")
        self.status_lbl.pack(anchor="w",pady=(4,0))

    def _logpane(self):
        tk.Frame(self,bg=C["shadow"],height=2).pack(fill="x",padx=16,pady=(4,0))
        lh=tk.Frame(self,bg=C["sky"]); lh.pack(fill="x",padx=16,pady=(4,0))
        tk.Label(lh,text="📜 Live Activity Log",font=("Comic Sans MS",10,"bold"),
                 bg=C["sky"],fg=C["purple"]).pack(side="left")
        tk.Button(lh,text="🧹 Clear",command=self._clrlog,
                  bg=C["shadow"],fg=C["text"],relief="flat",
                  font=("Comic Sans MS",8),padx=6,cursor="hand2").pack(side="right")
        self.log_box=ScrolledText(self,bg="#fafbff",fg=C["text"],
                                  font=("Consolas",9),relief="flat",bd=0,
                                  state="disabled",height=8)
        self.log_box.pack(fill="both",padx=16,pady=(4,12))
        self.log_box.tag_config("ok",foreground=C["teal"])
        self.log_box.tag_config("warn",foreground=C["orange"])
        self.log_box.tag_config("head",foreground=C["purple"])
        self.log_box.tag_config("star",foreground=C["pink"])
        self.log_box.tag_config("info",foreground=C["text"])

    # ── HELPERS ───────────────────────────────────────────────────────────────
    def _note(self,p,t,bg):
        tk.Label(p,text=t,font=("Arial",8),bg=bg,fg=C["muted"],
                 anchor="w",wraplength=460,justify="left").pack(anchor="w",pady=(0,4))

    def _textbox(self,p,content,bg,height):
        t=tk.Text(p,bg=C["white"],fg=C["text"],insertbackground=C["text"],
                  relief="flat",font=("Consolas",10),height=height,
                  padx=8,pady=6,wrap="word",
                  highlightthickness=2,highlightcolor=C["purple"],
                  selectbackground=C["purple"],selectforeground=C["white"])
        t.insert("1.0",content); t.pack(fill="x",pady=(0,4)); return t

    def _browse(self):
        d=filedialog.askdirectory(title="📂 Pick your save folder!")
        if d: self.out_dir_var.set(d)

    def _clrlog(self):
        self.log_box.config(state="normal"); self.log_box.delete("1.0","end"); self.log_box.config(state="disabled")

    def _log(self,tag,msg):
        self.log_box.config(state="normal"); self.log_box.insert("end",msg+"\n",tag)
        self.log_box.see("end"); self.log_box.config(state="disabled")

    def _status(self,msg,col=None):
        self.status_lbl.config(text=msg,fg=col or C["muted"])

    # ── START ─────────────────────────────────────────────────────────────────
    def _start(self):
        if self._running: return
        subs=[s.strip().lstrip("r/").strip()
              for s in self.sub_text.get("1.0","end").replace("\n",",").split(",") if s.strip()]
        if not subs:
            messagebox.showerror("Oops! 😅","Please enter at least one subreddit! 📋"); return
        kws=[k.strip() for k in self.kw_text.get("1.0","end").splitlines() if k.strip()]
        if not kws:
            messagebox.showerror("Oops! 😅","Please enter at least one keyword! 🔍"); return
        fname=self.fname_var.get().strip()
        if not fname.endswith(".xlsx"): fname+=".xlsx"
        out_dir=Path(self.out_dir_var.get().strip())
        out_path=str(out_dir/fname)
        config={"subreddits":subs,"keywords":kws,
                "max_posts":self.max_posts_var.get(),
                "max_comments":self.max_comments_var.get(),
                "days_back":self.days_back_var.get(),
                "sort_by":self.sort_var.get(),
                "time_filter":self.time_var.get()}
        self._running=True
        self.run_btn.update_label("⏳  Working… please wait!")
        self.run_btn.update_color(C["orange"])
        self.prog_var.set(0)
        self._status("🚀 Launching scrape rocket… hang tight!",C["purple"])
        threading.Thread(target=self._worker,args=(config,out_path,out_dir),daemon=True).start()

    def _worker(self,config,out_path,out_dir):
        try:
            out_dir.mkdir(parents=True,exist_ok=True)
            records=scrape(config,self._log_q,self._prog_q)
            if records:
                self._log_q.put(("info",f"\n💾 Saving to {out_path}…"))
                export_excel(records,out_path,config["max_comments"])
                self._log_q.put(("star",f"🎉 ALL DONE! {len(records)} posts saved to Excel!"))
                self._prog_q.put(("done",out_path,len(records)))
            else:
                self._log_q.put(("warn","😕 No posts found — try different keywords or more days back."))
                self._prog_q.put(("empty",))
        except Exception as e:
            self._log_q.put(("warn",f"💥 Error: {e}"))
            self._prog_q.put(("error",str(e)))

    # ── POLL ──────────────────────────────────────────────────────────────────
    def _poll(self):
        while not self._log_q.empty():
            tag,msg=self._log_q.get_nowait(); self._log(tag,msg)
        while not self._prog_q.empty():
            item=self._prog_q.get_nowait()
            if isinstance(item,float):
                self.prog_var.set(item); self._status(f"⚡ Progress: {int(item*100)}%  Keep going!",C["purple"])
            elif isinstance(item,tuple):
                code=item[0]
                self.run_btn.update_label("🚀  GENERATE REPORT!")
                self.run_btn.update_color(C["purple"])
                self._running=False
                if code=="done":
                    _,path,n=item; self.prog_var.set(1.0)
                    self._status(f"🎉 {n} posts saved → {path}",C["green"])
                    if messagebox.askyesno("🎉 Yahoo!",f"✅ {n} posts saved!\n\n📄 {path}\n\n👀 Open the output folder?"):
                        import subprocess,platform
                        folder=str(Path(path).parent)
                        if platform.system()=="Windows": os.startfile(folder)
                        elif platform.system()=="Darwin": subprocess.Popen(["open",folder])
                        else: subprocess.Popen(["xdg-open",folder])
                elif code=="empty":
                    self._status("😕 No data found. Try broader keywords or more days!",C["orange"])
                else:
                    self._status(f"💥 Error: {item[1]}",C["red"])
        self.after(200,self._poll)


if __name__=="__main__":
    app=FunScraperApp()
    app.mainloop()
