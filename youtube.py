"""
main.py  —  🎉 GenAI Ads YouTube Scraper  (No API Key! Fun Edition!)
═══════════════════════════════════════════════════════════════════════
Run:  python main.py

Auto-installs:  yt-dlp  openpyxl
No API key needed — uses yt-dlp to scrape YouTube directly!
"""

from __future__ import annotations

import sys, os, threading, queue, time, random, re, math, subprocess
from datetime import datetime, timedelta, timezone
from pathlib import Path

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.scrolledtext import ScrolledText

# ── auto-install deps ────────────────────────────────────────────────────────
def pip(pkg):
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])

try:
    import yt_dlp
except ImportError:
    print("Installing yt-dlp…"); pip("yt-dlp"); import yt_dlp

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl…"); pip("openpyxl")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
#  PALETTE
# ══════════════════════════════════════════════════════════════════════════════
C = {
    "sky":      "#fff0f0",
    "lavender": "#fff0f8",
    "mint":     "#f0fff8",
    "peach":    "#fff5ec",
    "lemon":    "#fffce8",
    "pink":     "#ff4444",
    "purple":   "#cc0000",
    "blue":     "#ff6b6b",
    "teal":     "#00c9a7",
    "orange":   "#ff8c42",
    "yellow":   "#ffd93d",
    "red":      "#cc0000",
    "green":    "#2ecc71",
    "text":     "#2d2d4e",
    "muted":    "#8888aa",
    "white":    "#ffffff",
    "shadow":   "#f0d0d0",
}
RAINBOW = [C["pink"], C["orange"], C["yellow"], C["green"], C["blue"], C["purple"]]

DEFAULT_CHANNELS = [
    "@MrBeast",
    "@mkbhd",
    "@LinusTechTips",
    "@TEDxTalks",
    "@Fireship",
]
DEFAULT_KEYWORDS = [
    "generative AI advertising",
    "AI ads",
    "GenAI marketing",
    "AI creative ads",
    "LLM advertising",
    "AI ad generation",
    "ChatGPT ads",
    "AI content marketing",
]
ORDER_OPTIONS = ["relevance", "date", "viewCount", "rating"]
COMMENT_COLOURS = [
    "FFF9C4","FFE0B2","F8BBD0","E1BEE7","C8E6C9",
    "B3E5FC","DCEDC8","F0F4C3","FFE0B2","FCE4EC",
]


# ══════════════════════════════════════════════════════════════════════════════
#  YT-DLP SCRAPER  (no API key)
# ══════════════════════════════════════════════════════════════════════════════

def _base_opts():
    return {
        "quiet":              True,
        "no_warnings":        True,
        "ignoreerrors":       True,
        "extract_flat":       True,
        "skip_download":      True,
        "nocheckcertificate": True,
    }


def _fmt_date(raw: str) -> str:
    if raw and len(raw) == 8 and raw.isdigit():
        return f"{raw[:4]}-{raw[4:6]}-{raw[6:]}"
    return raw or ""


def _parse_entry(entry: dict, channel_fallback: str) -> dict:
    vid_id = entry.get("id", "")
    return {
        "id":         vid_id,
        "channel":    entry.get("uploader") or entry.get("channel") or channel_fallback,
        "title":      (entry.get("title") or "").strip(),
        "url":        f"https://www.youtube.com/watch?v={vid_id}" if vid_id else "",
        "views":      int(entry.get("view_count")    or 0),
        "likes":      int(entry.get("like_count")    or 0),
        "n_comments": int(entry.get("comment_count") or 0),
        "date":       _fmt_date(entry.get("upload_date") or ""),
    }


def _get_channel_id(channel_input: str, log_q) -> str | None:
    """
    Resolve a channel name / @handle to a UC... channel ID using yt-dlp.
    This is robust — works for handles, slugs, full URLs.
    """
    s = channel_input.strip()
    if s.startswith("http"):
        url = s
    elif s.startswith("@"):
        url = f"https://www.youtube.com/{s}"
    else:
        url = f"https://www.youtube.com/@{s}"

    opts = {
        **_base_opts(),
        "extract_flat": True,
        "playlistend":  1,
    }
    try:
        with yt_dlp.YoutubeDL(opts) as ydl:
            info = ydl.extract_info(url, download=False)
            if info:
                cid = info.get("channel_id") or info.get("uploader_id") or ""
                if cid:
                    log_q.put(("ok", f"  ✅ Resolved '{channel_input}' → {cid}"))
                    return cid
    except Exception as e:
        log_q.put(("warn", f"  ⚠️  Could not resolve channel ID for '{channel_input}': {e}"))
    return None


def search_channel_videos(channel_input: str, keyword: str, max_results: int,
                           order: str, days_back: int, log_q) -> list:
    """
    Search videos within a channel using two strategies:
    1. ytsearch with channel ID filter (most reliable)
    2. ytsearch with channel name as fallback
    """
    results = []
    seen    = set()

    daterange = None
    if days_back > 0:
        start = (datetime.now() - timedelta(days=days_back)).strftime("%Y%m%d")
        try:
            daterange = yt_dlp.utils.DateRange(start, "99991231")
        except Exception:
            pass

    # ── Strategy 1: resolve channel ID → search with channelId filter ─────────
    channel_id = _get_channel_id(channel_input, log_q)
    if channel_id:
        try:
            import urllib.parse
            # Use YouTube's own search URL with channel filter
            search_url = (
                f"https://www.youtube.com/results?search_query="
                f"{urllib.parse.quote(keyword)}&sp=EgIQAQ%253D%253D"
            )
            # Better: search within channel videos playlist
            channel_search_url = f"https://www.youtube.com/channel/{channel_id}/search?query={urllib.parse.quote(keyword)}"
            opts = {
                **_base_opts(),
                "playlistend": max_results * 3,
            }
            if daterange:
                opts["daterange"] = daterange
            with yt_dlp.YoutubeDL(opts) as ydl:
                info = ydl.extract_info(channel_search_url, download=False)
                if info and info.get("entries"):
                    for entry in (info["entries"] or []):
                        if not entry or not entry.get("id"): continue
                        if entry["id"] in seen: continue
                        seen.add(entry["id"])
                        results.append(_parse_entry(entry, channel_input))
                        if len(results) >= max_results: break
        except Exception as e:
            log_q.put(("warn", f"  ⚠️  Channel ID search failed: {e}"))

    # ── Strategy 2: ytsearch for keyword + channel name, then filter ──────────
    if len(results) < max_results:
        if results:
            log_q.put(("info", f"  🔄  Supplementing with global search…"))
        else:
            log_q.put(("info", f"  ↩️  Using ytsearch for \"{keyword}\"…"))
        try:
            ch_clean = channel_input.lstrip("@").strip()
            n_fetch  = (max_results - len(results)) * 5
            query    = f"ytsearch{n_fetch}:{keyword} {ch_clean}"
            opts2    = {**_base_opts()}
            if daterange:
                opts2["daterange"] = daterange
            with yt_dlp.YoutubeDL(opts2) as ydl:
                info = ydl.extract_info(query, download=False)
                if info and info.get("entries"):
                    ch_lower = ch_clean.lower()
                    for entry in (info["entries"] or []):
                        if not entry or not entry.get("id"): continue
                        uploader = (entry.get("uploader") or entry.get("channel") or "").lower()
                        # match if channel name appears in uploader or vice-versa
                        if ch_lower in uploader or uploader in ch_lower or \
                           (channel_id and entry.get("channel_id") == channel_id):
                            if entry["id"] not in seen:
                                seen.add(entry["id"])
                                results.append(_parse_entry(entry, channel_input))
                        if len(results) >= max_results: break
        except Exception as e:
            log_q.put(("warn", f"  ⚠️  ytsearch failed: {e}"))

    return results[:max_results]


def fetch_video_details(video_url: str) -> dict:
    """Get full stats for a single video (views, likes, comment count, date)."""
    try:
        opts = {
            "quiet":             True,
            "no_warnings":       True,
            "ignoreerrors":      True,
            "skip_download":     True,
            "nocheckcertificate":True,
        }
        with yt_dlp.YoutubeDL(opts) as ydl:
            info = ydl.extract_info(video_url, download=False)
            if info:
                return {
                    "views":      int(info.get("view_count")    or 0),
                    "likes":      int(info.get("like_count")    or 0),
                    "n_comments": int(info.get("comment_count") or 0),
                    "date":       _fmt_date(info.get("upload_date") or ""),
                    "channel":    info.get("uploader") or info.get("channel") or "",
                }
    except Exception:
        pass
    return {}


def fetch_top_comments(video_url: str, n: int, log_q) -> list:
    """Fetch top N comments using yt-dlp's comment extraction."""
    if n <= 0:
        return []
    try:
        opts = {
            "quiet":          True,
            "no_warnings":    True,
            "ignoreerrors":   True,
            "skip_download":  True,
            "nocheckcertificate": True,
            "getcomments":    True,
            "extractor_args": {
                "youtube": {
                    "max_comments":        [str(n * 3)],
                    "max_comment_depth":   ["1"],
                }
            },
        }
        with yt_dlp.YoutubeDL(opts) as ydl:
            info = ydl.extract_info(video_url, download=False)
            if not info:
                return []
            raw = info.get("comments") or []
            # filter top-level only, sort by likes
            top_level = [c for c in raw if not c.get("parent") or c.get("parent") == "root"]
            top_level.sort(key=lambda c: int(c.get("like_count") or 0), reverse=True)
            comments = []
            for c in top_level:
                text = re.sub(r"<[^>]+>", "", c.get("text") or "")
                text = re.sub(r"\s+", " ", text).strip()[:500]
                if text:
                    comments.append(text)
                if len(comments) >= n:
                    break
            return comments
    except Exception as e:
        log_q.put(("warn", f"  ⚠️  Comments failed: {e}"))
        return []


def scrape(config: dict, log_q, progress_q) -> list:
    records, seen_ids = [], set()
    channels = config["channels"]

    for ch_i, channel_input in enumerate(channels, 1):
        log_q.put(("head", f"🚀 Diving into: {channel_input}  ({ch_i}/{len(channels)})…"))
        ch_count = 0

        for keyword in config["keywords"]:
            if ch_count >= config["max_videos"]:
                break

            log_q.put(("info", f"  🔍 Keyword: \"{keyword}\"…"))
            videos = search_channel_videos(
                channel_input, keyword,
                config["max_videos"] - ch_count,
                config["order_by"],
                config["days_back"],
                log_q,
            )
            time.sleep(random.uniform(1.0, 2.5))

            for vid in videos:
                if ch_count >= config["max_videos"]: break
                if not vid["id"] or vid["id"] in seen_ids: continue
                seen_ids.add(vid["id"])

                # Always enrich with full video page (accurate stats + enables comments)
                log_q.put(("info", f"     📊 Loading video details…"))
                d = fetch_video_details(vid["url"])
                if d:
                    vid.update({k: v for k, v in d.items() if v})
                time.sleep(random.uniform(0.8, 1.5))

                log_q.put(("ok",
                    f"  🎬 {vid['title'][:55]}…  "
                    f"👁{vid['views']:,}  👍{vid['likes']:,}  💬{vid['n_comments']:,}"))

                comments = []
                if config["max_comments"] > 0:
                    log_q.put(("info", "     💬 Fetching comments…"))
                    comments = fetch_top_comments(vid["url"], config["max_comments"], log_q)
                    log_q.put(("info", f"     ✅ Got {len(comments)} comment(s)"))
                    time.sleep(random.uniform(1.0, 2.0))

                record = {
                    "channel":     vid["channel"] or channel_input,
                    "video_title": vid["title"],
                    "video_url":   vid["url"],
                    "views":       vid["views"],
                    "likes":       vid["likes"],
                    "date":        vid["date"],
                }
                for i in range(1, config["max_comments"] + 1):
                    record[f"comment_{i}"] = comments[i-1] if i <= len(comments) else ""

                records.append(record)
                ch_count += 1

        log_q.put(("ok", f"  🎯 {channel_input}: {ch_count} videos collected!"))
        progress_q.put(ch_i / len(channels))

    log_q.put(("star", f"🎉 WOW! Total videos scraped: {len(records)}"))
    return records


# ══════════════════════════════════════════════════════════════════════════════
#  EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════
def export_excel(records: list, filepath: str, max_comments: int):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GenAI Ads — YouTube"

    HF = PatternFill("solid", fgColor="C00000")
    AF = PatternFill("solid", fgColor="FFF2F2")
    HN = Font(bold=True, color="FFFFFF", size=11)
    LN = Font(color="C00000", underline="single", size=10)
    NF = Font(size=10)
    CE = Alignment(horizontal="center", vertical="top", wrap_text=False)
    LW = Alignment(horizontal="left",   vertical="top", wrap_text=True)
    TN = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    cols = ["Channel","Video Title","Video Link","Views","Likes","Date"] + \
           [f"Comment {i}" for i in range(1, max_comments+1)]
    ws.append(cols)
    for ci in range(1, len(cols)+1):
        c = ws.cell(row=1, column=ci)
        c.fill=HF; c.font=HN; c.alignment=CE; c.border=TN
    ws.row_dimensions[1].height = 28

    for ri, rec in enumerate(records, start=2):
        rf = AF if ri%2==0 else PatternFill("solid", fgColor="FFFFFF")

        def cell(col, val, fill=None, font=None, align=None):
            c = ws.cell(row=ri, column=col, value=val)
            c.fill=fill or rf; c.font=font or NF; c.alignment=align or CE; c.border=TN
            return c

        cell(1, rec["channel"], fill=PatternFill("solid", fgColor="FFD6D6"),
             font=Font(bold=True, size=10))
        cell(2, rec["video_title"], align=LW)
        lc = cell(3, "▶ Watch Video", font=LN)
        lc.hyperlink = rec["video_url"]
        cell(4, rec["views"])
        cell(5, rec["likes"])
        cell(6, rec["date"])
        for i in range(1, max_comments+1):
            txt = rec.get(f"comment_{i}", "")
            col = COMMENT_COLOURS[(i-1) % len(COMMENT_COLOURS)]
            cell(6+i, txt,
                 fill=PatternFill("solid", fgColor=col) if txt else rf,
                 align=LW)
        ws.row_dimensions[ri].height = 60

    widths = [20,45,14,12,10,12] + [35]*max_comments
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)


# ══════════════════════════════════════════════════════════════════════════════
#  UI WIDGETS
# ══════════════════════════════════════════════════════════════════════════════
class SparkleCanvas(tk.Canvas):
    SYMBOLS = ["▶","★","✦","✧","◆","✿","❋","♦","●"]
    def __init__(self, parent, **kw):
        super().__init__(parent, highlightthickness=0, **kw)
        self._stars = []; self.after(200, self._init)
    def _init(self):
        w = max(self.winfo_width(), 800); h = max(self.winfo_height(), 100)
        for _ in range(28):
            self._stars.append({
                "x": random.uniform(0,w), "y": random.uniform(0,h),
                "vy": random.uniform(-0.5,-0.15), "vx": random.uniform(-0.2,0.2),
                "size": random.randint(9,20), "sym": random.choice(self.SYMBOLS),
                "color": random.choice(RAINBOW), "phase": random.uniform(0,6.28),
                "dphase": random.uniform(0.03,0.09),
            })
        self._tick()
    def _tick(self):
        self.delete("sp")
        w = max(self.winfo_width(), 800); h = max(self.winfo_height(), 100)
        for s in self._stars:
            s["x"] = (s["x"]+s["vx"]) % w; s["y"] += s["vy"]
            if s["y"] < -24: s["y"] = h+10
            s["phase"] += s["dphase"]
            sz = s["size"] + int(3*math.sin(s["phase"]))
            self.create_text(s["x"], s["y"], text=s["sym"],
                             font=("Arial",sz), fill=s["color"], tags="sp")
        self.after(40, self._tick)


class PillButton(tk.Canvas):
    def __init__(self, parent, text, command, color, **kw):
        super().__init__(parent, highlightthickness=0, cursor="hand2",
                         bg=parent.cget("bg"), **kw)
        self._text=text; self._cmd=command; self._color=color
        self._phase=0.0; self._pressed=False
        self.bind("<Button-1>", self._press)
        self.bind("<ButtonRelease-1>", self._release)
        self.bind("<Configure>", lambda _: self._draw())
        self.after(20, self._animate)
    def _lgt(self,h,a=55):
        h=h.lstrip("#"); r,g,b=int(h[0:2],16),int(h[2:4],16),int(h[4:6],16)
        return f"#{min(255,r+a):02x}{min(255,g+a):02x}{min(255,b+a):02x}"
    def _drk(self,h,a=30):
        h=h.lstrip("#"); r,g,b=int(h[0:2],16),int(h[2:4],16),int(h[4:6],16)
        return f"#{max(0,r-a):02x}{max(0,g-a):02x}{max(0,b-a):02x}"
    def _pill(self,x1,y1,x2,y2,r,**kw):
        self.create_arc(x1,y1,x1+2*r,y1+2*r,start=90,extent=90,style="pieslice",**kw)
        self.create_arc(x2-2*r,y1,x2,y1+2*r,start=0,extent=90,style="pieslice",**kw)
        self.create_arc(x1,y2-2*r,x1+2*r,y2,start=180,extent=90,style="pieslice",**kw)
        self.create_arc(x2-2*r,y2-2*r,x2,y2,start=270,extent=90,style="pieslice",**kw)
        self.create_rectangle(x1+r,y1,x2-r,y2,**kw)
        self.create_rectangle(x1,y1+r,x2,y2-r,**kw)
    def _draw(self):
        self.delete("all"); w,h=self.winfo_width(),self.winfo_height()
        if w<10 or h<10: return
        g=int(5+4*math.sin(self._phase)); r=h//2
        fill=self._drk(self._color) if self._pressed else self._color
        self._pill(g,g,w-2,h-2,r,fill="#c0bce8",outline="")
        self._pill(0,0,w-g-2,h-g-2,r,fill=fill,outline="")
        self._pill(8,5,w-g-10,h//2-2,r-4,fill=self._lgt(fill,70),outline="")
        self.create_text((w-g)//2,(h-g)//2,text=self._text,
                         font=("Comic Sans MS",14,"bold"),fill="#ffffff")
    def _animate(self): self._phase+=0.06; self._draw(); self.after(30,self._animate)
    def _press(self,_): self._pressed=True; self._cmd(); self._draw()
    def _release(self,_): self._pressed=False; self._draw()
    def update_label(self,t): self._text=t; self._draw()
    def update_color(self,c): self._color=c; self._draw()


def attach_tip(widget, text):
    tip=[None]
    def show(_):
        if tip[0]: return
        w=tk.Toplevel(widget); w.wm_overrideredirect(True)
        x=widget.winfo_rootx()+22; y=widget.winfo_rooty()+widget.winfo_height()+4
        w.wm_geometry(f"+{x}+{y}")
        tk.Label(w,text=text,bg="#fffde7",fg=C["text"],font=("Arial",9),
                 relief="solid",bd=1,padx=10,pady=6,wraplength=300).pack()
        tip[0]=w
    def hide(_):
        if tip[0]: tip[0].destroy(); tip[0]=None
    widget.bind("<Enter>",show); widget.bind("<Leave>",hide)


def make_card(parent, bg, emoji, title, subtitle, pady=10):
    outer=tk.Frame(parent,bg=bg,relief="flat"); outer.pack(fill="x",pady=(0,pady))
    inner=tk.Frame(outer,bg=bg); inner.pack(fill="x",padx=14,pady=12)
    hdr=tk.Frame(inner,bg=bg); hdr.pack(fill="x",pady=(0,4))
    tk.Label(hdr,text=emoji,font=("Arial",18),bg=bg).pack(side="left",padx=(0,6))
    tk.Label(hdr,text=title,font=("Comic Sans MS",12,"bold"),bg=bg,fg=C["text"]).pack(side="left")
    if subtitle:
        tk.Label(hdr,text=f"  — {subtitle}",font=("Arial",8),bg=bg,fg=C["muted"]).pack(side="left")
    return inner


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN APP
# ══════════════════════════════════════════════════════════════════════════════
class FunScraperApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("🎉 GenAI Ads YouTube Scraper — No API Key!")
        self.geometry("1090x940"); self.minsize(920,760)
        self.configure(bg=C["sky"])
        self.update_idletasks()
        sw,sh=self.winfo_screenwidth(),self.winfo_screenheight()
        self.geometry(f"1090x940+{(sw-1090)//2}+{(sh-940)//2}")
        self._log_q=queue.Queue(); self._prog_q=queue.Queue(); self._running=False
        self._build(); self._poll()

    def _build(self): self._hdr(); self._body(); self._logpane()

    def _hdr(self):
        hdr=tk.Frame(self,bg=C["sky"],height=96); hdr.pack(fill="x"); hdr.pack_propagate(False)
        sc=SparkleCanvas(hdr,bg=C["sky"]); sc.place(x=0,y=0,relwidth=1,relheight=1)
        tf=tk.Frame(hdr,bg=C["sky"]); tf.place(relx=0.5,rely=0.5,anchor="center")
        row=tk.Frame(tf,bg=C["sky"]); row.pack()
        for txt,col in [("▶️",C["red"]),(" GenAI ",C["purple"]),("Ads ",C["pink"]),
                        ("YouTube ",C["red"]),("Scraper ","#ff8c42"),("🎬",C["green"])]:
            tk.Label(row,text=txt,font=("Comic Sans MS",26,"bold"),bg=C["sky"],fg=col).pack(side="left")
        tk.Label(tf,text="✨  No API Key needed!  Enter channels + keywords, then hit GENERATE REPORT! ✨",
                 font=("Comic Sans MS",10),bg=C["sky"],fg=C["muted"]).pack()

    def _body(self):
        outer=tk.Frame(self,bg=C["sky"]); outer.pack(fill="both",expand=True,padx=16,pady=(6,0))
        cv=tk.Canvas(outer,bg=C["sky"],highlightthickness=0)
        vsb=ttk.Scrollbar(outer,orient="vertical",command=cv.yview)
        cv.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right",fill="y"); cv.pack(side="left",fill="both",expand=True)
        sf=tk.Frame(cv,bg=C["sky"]); fid=cv.create_window((0,0),window=sf,anchor="nw")
        sf.bind("<Configure>",lambda _:cv.configure(scrollregion=cv.bbox("all")))
        cv.bind("<Configure>",lambda e:cv.itemconfig(fid,width=e.width))
        cv.bind_all("<MouseWheel>",lambda e:cv.yview_scroll(int(-1*(e.delta/120)),"units"))
        self._form(sf)

    def _form(self, parent):
        cols=tk.Frame(parent,bg=C["sky"]); cols.pack(fill="x",pady=(4,0))
        left=tk.Frame(cols,bg=C["sky"]); right=tk.Frame(cols,bg=C["sky"])
        left.pack(side="left",fill="both",expand=True,padx=(0,8))
        right.pack(side="left",fill="both",expand=True)

        # ── LEFT: Channels ────────────────────────────────────────────────────
        ci=make_card(left,C["lavender"],"📺","YouTube Channels","one per line — name or @handle")
        self._note(ci,
            "Enter channel names or @handles, one per line.\n"
            "Examples:  @MrBeast   Linus Tech Tips   @mkbhd",
            C["lavender"])
        self.ch_text=self._textbox(ci,"\n".join(DEFAULT_CHANNELS),C["lavender"],5)

        # ── LEFT: Keywords ────────────────────────────────────────────────────
        ki=make_card(left,C["mint"],"🔍","Search Keywords","one keyword or phrase per line")
        self._note(ki,
            "Each keyword is searched within every channel above.\n"
            "More keywords = more videos found (but takes longer).",
            C["mint"])
        self.kw_text=self._textbox(ki,"\n".join(DEFAULT_KEYWORDS),C["mint"],7)

        # ── RIGHT: Settings ───────────────────────────────────────────────────
        pi=make_card(right,C["peach"],"⚙️","Scrape Settings","control how much data to collect")
        g=tk.Frame(pi,bg=C["peach"]); g.pack(fill="x"); g.columnconfigure(1,weight=1)

        self.max_videos_var   = tk.IntVar(value=20)
        self.max_comments_var = tk.IntVar(value=5)
        self.days_back_var    = tk.IntVar(value=180)
        self.order_var        = tk.StringVar(value="relevance")

        spinfields=[
            ("🎬 Max Videos / Channel", self.max_videos_var, 1, 200,
             "How many videos per channel?\n20 is a great start. More = slower."),
            ("💬 Comments per Video",   self.max_comments_var, 0, 20,
             "Each comment = its own Excel column.\n0 = skip comments (much faster!)."),
            ("📅 Days Back",            self.days_back_var, 0, 3650,
             "Only grab videos from last N days.\n0 = no date filter."),
        ]
        for rn,(lbl,var,lo,hi,tip) in enumerate(spinfields):
            l=tk.Label(g,text=lbl,font=("Comic Sans MS",10,"bold"),bg=C["peach"],fg=C["text"],anchor="w")
            l.grid(row=rn*2,column=0,sticky="w",pady=(10,0)); attach_tip(l,tip)
            sb=tk.Spinbox(g,textvariable=var,from_=lo,to=hi,width=8,
                          bg=C["white"],fg=C["text"],relief="flat",
                          font=("Comic Sans MS",11,"bold"),buttonbackground=C["orange"],
                          highlightthickness=2,highlightcolor=C["orange"],insertbackground=C["text"])
            sb.grid(row=rn*2,column=1,sticky="w",pady=(10,0),padx=(14,0))
            tk.Label(g,text=tip.split("\n")[0],font=("Arial",8),
                     bg=C["peach"],fg=C["muted"],anchor="w").grid(
                         row=rn*2+1,column=0,columnspan=2,sticky="w",padx=4)

        rn=len(spinfields)
        l=tk.Label(g,text="🔀 Sort Results By",font=("Comic Sans MS",10,"bold"),
                   bg=C["peach"],fg=C["text"],anchor="w")
        l.grid(row=rn*2,column=0,sticky="w",pady=(10,0))
        attach_tip(l,"relevance=best match  date=newest  viewCount=most views  rating=top rated")
        dd=ttk.Combobox(g,textvariable=self.order_var,values=ORDER_OPTIONS,
                        state="readonly",width=14,font=("Comic Sans MS",10))
        dd.grid(row=rn*2,column=1,sticky="w",pady=(10,0),padx=(14,0))
        tk.Label(g,text="How to sort search results",font=("Arial",8),
                 bg=C["peach"],fg=C["muted"],anchor="w").grid(
                     row=rn*2+1,column=0,columnspan=2,sticky="w",padx=4)

        tk.Label(pi,
            text="⏱️  Tip: Comments are slow to fetch. Set Comments = 0 for a fast first run!",
            font=("Arial",8,"italic"),bg=C["peach"],fg=C["muted"],
            wraplength=400,justify="left",anchor="w").pack(anchor="w",pady=(8,0))

        # ── RIGHT: Output ─────────────────────────────────────────────────────
        oi=make_card(right,C["lemon"],"💾","Output File","where to save the Excel report")
        self._note(oi,"📁 Folder to save in  (click Browse or type a path)",C["lemon"])
        pr=tk.Frame(oi,bg=C["lemon"]); pr.pack(fill="x",pady=(4,0))
        self.out_dir_var=tk.StringVar(value=str(Path.cwd()/"data"))
        pe=tk.Entry(pr,textvariable=self.out_dir_var,bg=C["white"],fg=C["text"],
                    relief="flat",font=("Comic Sans MS",10),bd=0,
                    highlightthickness=2,highlightcolor=C["yellow"],insertbackground=C["text"])
        pe.pack(side="left",fill="x",expand=True,ipady=6,ipadx=6)
        attach_tip(pe,"Folder where the Excel file will be saved.\nCreated automatically if it doesn't exist.")
        tk.Button(pr,text="📂 Browse…",command=self._browse,
                  bg=C["orange"],fg=C["white"],relief="flat",
                  font=("Comic Sans MS",10,"bold"),
                  activebackground=C["red"],activeforeground=C["white"],
                  padx=8,pady=4,cursor="hand2",bd=0).pack(side="left",padx=(6,0))
        self._note(oi,"📄 File name  (must end with .xlsx)",C["lemon"])
        self.fname_var=tk.StringVar(value="genai_ads_youtube.xlsx")
        fe=tk.Entry(oi,textvariable=self.fname_var,bg=C["white"],fg=C["text"],
                    relief="flat",font=("Comic Sans MS",10),bd=0,
                    highlightthickness=2,highlightcolor=C["yellow"],insertbackground=C["text"])
        fe.pack(fill="x",ipady=6,ipadx=6,pady=(4,0))
        attach_tip(fe,"Name for your Excel report file. Must end with .xlsx")

        # ── BIG BUTTON ────────────────────────────────────────────────────────
        bf=tk.Frame(right,bg=C["sky"]); bf.pack(fill="x",pady=(8,0))
        self.run_btn=PillButton(bf,"▶  GENERATE REPORT!",self._start,C["red"],height=60)
        self.run_btn.pack(fill="x")

        pf=tk.Frame(right,bg=C["sky"]); pf.pack(fill="x",pady=(8,0))
        self.prog_var=tk.DoubleVar(value=0)
        sty=ttk.Style(); sty.theme_use("clam")
        sty.configure("Candy.Horizontal.TProgressbar",
                      troughcolor=C["shadow"],background=C["red"],
                      lightcolor=C["red"],darkcolor=C["red"],
                      bordercolor=C["sky"],thickness=16)
        ttk.Progressbar(pf,variable=self.prog_var,maximum=1.0,
                        style="Candy.Horizontal.TProgressbar").pack(fill="x")
        self.status_lbl=tk.Label(pf,
            text="👋 Ready!  Fill in channels + keywords, then hit the big red button!",
            bg=C["sky"],fg=C["muted"],font=("Comic Sans MS",9),wraplength=500,justify="left")
        self.status_lbl.pack(anchor="w",pady=(4,0))

    def _logpane(self):
        tk.Frame(self,bg=C["shadow"],height=2).pack(fill="x",padx=16,pady=(4,0))
        lh=tk.Frame(self,bg=C["sky"]); lh.pack(fill="x",padx=16,pady=(4,0))
        tk.Label(lh,text="📜 Live Activity Log",font=("Comic Sans MS",10,"bold"),
                 bg=C["sky"],fg=C["red"]).pack(side="left")
        tk.Button(lh,text="🧹 Clear",command=self._clrlog,
                  bg=C["shadow"],fg=C["text"],relief="flat",
                  font=("Comic Sans MS",8),padx=6,cursor="hand2").pack(side="right")
        self.log_box=ScrolledText(self,bg="#fff8f8",fg=C["text"],
                                  font=("Consolas",9),relief="flat",bd=0,
                                  state="disabled",height=9)
        self.log_box.pack(fill="both",padx=16,pady=(4,12))
        self.log_box.tag_config("ok",   foreground=C["teal"])
        self.log_box.tag_config("warn", foreground=C["orange"])
        self.log_box.tag_config("head", foreground=C["red"])
        self.log_box.tag_config("star", foreground=C["purple"])
        self.log_box.tag_config("info", foreground=C["text"])

    # ── helpers ───────────────────────────────────────────────────────────────
    def _note(self,p,t,bg):
        tk.Label(p,text=t,font=("Arial",8),bg=bg,fg=C["muted"],
                 anchor="w",wraplength=460,justify="left").pack(anchor="w",pady=(0,4))

    def _textbox(self,p,content,bg,height):
        t=tk.Text(p,bg=C["white"],fg=C["text"],insertbackground=C["text"],
                  relief="flat",font=("Consolas",10),height=height,
                  padx=8,pady=6,wrap="word",
                  highlightthickness=2,highlightcolor=C["red"],
                  selectbackground=C["red"],selectforeground=C["white"])
        t.insert("1.0",content); t.pack(fill="x",pady=(0,4)); return t

    def _browse(self):
        d=filedialog.askdirectory(title="📂 Pick your save folder!")
        if d: self.out_dir_var.set(d)

    def _clrlog(self):
        self.log_box.config(state="normal"); self.log_box.delete("1.0","end")
        self.log_box.config(state="disabled")

    def _log(self,tag,msg):
        self.log_box.config(state="normal"); self.log_box.insert("end",msg+"\n",tag)
        self.log_box.see("end"); self.log_box.config(state="disabled")

    def _status(self,msg,col=None):
        self.status_lbl.config(text=msg,fg=col or C["muted"])

    # ── start / worker / poll ─────────────────────────────────────────────────
    def _start(self):
        if self._running: return
        channels=[c.strip() for c in self.ch_text.get("1.0","end").splitlines() if c.strip()]
        if not channels:
            messagebox.showerror("Oops! 😅","Please enter at least one channel! 📺"); return
        kws=[k.strip() for k in self.kw_text.get("1.0","end").splitlines() if k.strip()]
        if not kws:
            messagebox.showerror("Oops! 😅","Please enter at least one keyword! 🔍"); return
        fname=self.fname_var.get().strip()
        if not fname.endswith(".xlsx"): fname+=".xlsx"
        out_dir=Path(self.out_dir_var.get().strip())
        out_path=str(out_dir/fname)
        config={
            "channels":     channels,
            "keywords":     kws,
            "max_videos":   self.max_videos_var.get(),
            "max_comments": self.max_comments_var.get(),
            "days_back":    self.days_back_var.get(),
            "order_by":     self.order_var.get(),
        }
        self._running=True
        self.run_btn.update_label("⏳  Working… please wait!")
        self.run_btn.update_color(C["orange"])
        self.prog_var.set(0)
        self._status("🚀 Launching scrape rocket… hang tight!",C["purple"])
        threading.Thread(target=self._worker,args=(config,out_path,out_dir),daemon=True).start()

    def _worker(self,config,out_path,out_dir):
        try:
            out_dir.mkdir(parents=True,exist_ok=True)
            records=scrape(config,self._log_q,self._prog_q)
            if records:
                self._log_q.put(("info",f"\n💾 Saving to {out_path}…"))
                export_excel(records,out_path,config["max_comments"])
                self._log_q.put(("star",f"🎉 ALL DONE! {len(records)} videos saved to Excel!"))
                self._prog_q.put(("done",out_path,len(records)))
            else:
                self._log_q.put(("warn","😕 No videos found — try different keywords, channels, or more days."))
                self._prog_q.put(("empty",))
        except Exception as e:
            import traceback
            self._log_q.put(("warn",f"💥 Error: {e}\n{traceback.format_exc()}"))
            self._prog_q.put(("error",str(e)))

    def _poll(self):
        while not self._log_q.empty():
            tag,msg=self._log_q.get_nowait(); self._log(tag,msg)
        while not self._prog_q.empty():
            item=self._prog_q.get_nowait()
            if isinstance(item,float):
                self.prog_var.set(item)
                self._status(f"⚡ Progress: {int(item*100)}%  Keep going!",C["purple"])
            elif isinstance(item,tuple):
                code=item[0]
                self.run_btn.update_label("▶  GENERATE REPORT!")
                self.run_btn.update_color(C["red"])
                self._running=False
                if code=="done":
                    _,path,n=item; self.prog_var.set(1.0)
                    self._status(f"🎉 {n} videos saved → {path}",C["green"])
                    if messagebox.askyesno("🎉 Yahoo!",
                                           f"✅ {n} videos saved!\n\n📄 {path}\n\n👀 Open the output folder?"):
                        import platform
                        folder=str(Path(path).parent)
                        if platform.system()=="Windows":   os.startfile(folder)
                        elif platform.system()=="Darwin":  subprocess.Popen(["open",folder])
                        else:                              subprocess.Popen(["xdg-open",folder])
                elif code=="empty":
                    self._status("😕 No data. Try broader keywords or more days!",C["orange"])
                else:
                    self._status(f"💥 Error: {item[1]}",C["red"])
        self.after(200,self._poll)


if __name__ == "__main__":
    app = FunScraperApp()
    app.mainloop()
