"""
main.py  —  🎉 GenAI Ads Reddit Scraper  (Fun Edition!)
══════════════════════════════════════════════════════
Run:  python main.py
"""

from __future__ import annotations

import sys, os, threading, queue, time, random, re, math
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Dict, List

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkinter.scrolledtext import ScrolledText

try:
    import requests
except ImportError:
    os.system(f"{sys.executable} -m pip install requests -q"); import requests

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
#  CANDY PALETTE
# ══════════════════════════════════════════════════════════════════════════════
C = {
    "sky":      "#eef6ff",
    "cloud":    "#ffffff",
    "lavender": "#f3eeff",
    "mint":     "#e5fbf4",
    "peach":    "#fff3ec",
    "lemon":    "#fffce8",
    "pink":     "#ff6eb4",
    "purple":   "#9b59ff",
    "blue":     "#3da9fc",
    "teal":     "#00c9a7",
    "orange":   "#ff8c42",
    "yellow":   "#ffd93d",
    "red":      "#ff4757",
    "green":    "#2ecc71",
    "text":     "#2d2d4e",
    "muted":    "#8888aa",
    "white":    "#ffffff",
    "shadow":   "#d8d8f0",
}
RAINBOW = [C["pink"], C["orange"], C["yellow"], C["green"], C["blue"], C["purple"]]

DEFAULT_SUBREDDITS = [
    "advertising","SocialMediaMarketing","FacebookAds","PPC",
    "digital_marketing","GoogleAds","marketing","ArtificialIntelligence",
]
DEFAULT_KEYWORDS = [
    "generative AI advertising","AI ads","GenAI marketing",
    "AI creative ads","LLM advertising","AI ad generation",
    "ChatGPT ads","AI content marketing",
]
SORT_OPTIONS = ["relevance","new","top","comments"]
TIME_OPTIONS = ["all","year","month","week","day"]
COMMENT_COLOURS = ["FFF9C4","FFE0B2","F8BBD0","E1BEE7","C8E6C9","B3E5FC","DCEDC8","F0F4C3","FFE0B2","FCE4EC"]


# ══════════════════════════════════════════════════════════════════════════════
#  SCRAPER
# ══════════════════════════════════════════════════════════════════════════════
SEARCH_URL = "https://www.reddit.com/r/{sub}/search.json"

def make_session():
    s = requests.Session()
    s.headers.update({"User-Agent":"Mozilla/5.0 (compatible; GenAI-Ads-Scraper/1.0; research)","Accept":"application/json"})
    return s

def search_subreddit(session, sub, query, sort, time_filter, limit, log_q):
    params={"q":query,"sort":sort,"restrict_sr":"1","t":time_filter,"limit":min(limit,100),"raw_json":"1"}
    try:
        r=session.get(SEARCH_URL.format(sub=sub),params=params,timeout=20)
        if r.status_code==429:
            wait=int(r.headers.get("Retry-After",30))
            log_q.put(("warn",f"⏳ Rate limited — waiting {wait}s…"))
            time.sleep(wait); return []
        if r.status_code!=200:
            log_q.put(("warn",f"😬 r/{sub} HTTP {r.status_code}")); return []
        return r.json().get("data",{}).get("children",[])
    except Exception as e:
        log_q.put(("warn",f"😬 r/{sub} error: {e}")); return []

def fetch_top_comments(session, permalink, n, log_q):
    url=f"https://www.reddit.com{permalink}.json?limit={n+5}&sort=top&raw_json=1"
    try:
        r=session.get(url,timeout=20)
        if r.status_code!=200: return []
        children=r.json()[1]["data"]["children"]
        comments=[]
        for child in children:
            body=child.get("data",{}).get("body","")
            if child.get("kind")=="t1" and body and body not in("[deleted]","[removed]"):
                comments.append(re.sub(r"\s+"," ",body).strip()[:500])
            if len(comments)>=n: break
        return comments
    except: return []

def scrape(config, log_q, progress_q, stop_event):
    """stop_event: threading.Event — set it to abort mid-scrape."""
    # Date range: use explicit start/end dates
    date_from = config["date_from"]   # datetime or None
    date_to   = config["date_to"]     # datetime or None

    records, seen_ids, session = [], set(), make_session()
    subs = config["subreddits"]

    for sub_i, sub in enumerate(subs, 1):
        if stop_event.is_set():
            log_q.put(("warn", "🛑 Stopped by user!"))
            break

        log_q.put(("head", f"🚀 Blasting into r/{sub}  ({sub_i}/{len(subs)})…"))
        sub_count = 0

        for keyword in config["keywords"]:
            if sub_count >= config["max_posts"] or stop_event.is_set():
                break
            children = search_subreddit(session, sub, keyword,
                                        config["sort_by"], config["time_filter"], 25, log_q)
            # Interruptible sleep
            for _ in range(int(random.uniform(12, 25))):
                if stop_event.is_set(): break
                time.sleep(0.1)

            for child in children:
                if sub_count >= config["max_posts"] or stop_event.is_set():
                    break
                post = child.get("data", {})
                pid  = post.get("id", "")
                if not pid or pid in seen_ids: continue
                seen_ids.add(pid)

                ts = post.get("created_utc", 0)
                dt = datetime.utcfromtimestamp(ts) if ts else None

                # Apply date range filter
                if dt:
                    if date_from and dt < date_from: continue
                    if date_to   and dt > date_to:   continue

                title      = post.get("title", "").strip()
                permalink  = post.get("permalink", "")
                score      = int(post.get("score", 0))
                n_comments = int(post.get("num_comments", 0))

                log_q.put(("ok", f"  ✨ {title[:55]}…  👍{score}  💬{n_comments}"))

                comments = []
                if permalink and n_comments > 0 and not stop_event.is_set():
                    comments = fetch_top_comments(session, permalink, config["max_comments"], log_q)
                    for _ in range(int(random.uniform(10, 20))):
                        if stop_event.is_set(): break
                        time.sleep(0.1)

                record = {"subreddit": f"r/{sub}", "post_title": title,
                          "post_url": f"https://www.reddit.com{permalink}",
                          "score": score, "date": dt.strftime("%Y-%m-%d") if dt else ""}
                for i in range(1, config["max_comments"]+1):
                    record[f"comment_{i}"] = comments[i-1] if i <= len(comments) else ""
                records.append(record)
                sub_count += 1

        if not stop_event.is_set():
            log_q.put(("ok", f"  🎯 r/{sub}: {sub_count} posts collected!"))
        progress_q.put(sub_i / len(subs))

    if stop_event.is_set():
        log_q.put(("warn", f"⚠️  Stopped early — {len(records)} posts collected so far."))
    else:
        log_q.put(("star", f"🎉 WOW! Total posts scraped: {len(records)}"))
    return records

def export_excel(records, filepath, max_comments):
    wb=openpyxl.Workbook(); ws=wb.active; ws.title="GenAI Ads — Reddit"
    HF=PatternFill("solid",fgColor="1F4E79"); AF=PatternFill("solid",fgColor="F2F7FC")
    HN=Font(bold=True,color="FFFFFF",size=11); LN=Font(color="0563C1",underline="single",size=10)
    NF=Font(size=10)
    CE=Alignment(horizontal="center",vertical="top",wrap_text=False)
    LW=Alignment(horizontal="left",vertical="top",wrap_text=True)
    TN=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    cols=["Subreddit","Post Title","Post Link","Score","Date"]+[f"Comment {i}" for i in range(1,max_comments+1)]
    ws.append(cols)
    for ci in range(1,len(cols)+1):
        c=ws.cell(row=1,column=ci); c.fill=HF; c.font=HN; c.alignment=CE; c.border=TN
    ws.row_dimensions[1].height=28
    for ri,rec in enumerate(records,start=2):
        rf=AF if ri%2==0 else PatternFill("solid",fgColor="FFFFFF")
        def cell(col,val,fill=None,font=None,align=None):
            c=ws.cell(row=ri,column=col,value=val)
            c.fill=fill or rf; c.font=font or NF; c.alignment=align or CE; c.border=TN
            return c
        cell(1,rec["subreddit"],fill=PatternFill("solid",fgColor="D6E4F0"),font=Font(bold=True,size=10))
        cell(2,rec["post_title"],align=LW)
        c=cell(3,"🔗 Open Post",font=LN); c.hyperlink=rec["post_url"]
        cell(4,rec["score"])
        cell(5,rec["date"])
        for i in range(1,max_comments+1):
            txt=rec.get(f"comment_{i}",""); clr=COMMENT_COLOURS[(i-1)%len(COMMENT_COLOURS)]
            cell(5+i,txt,fill=PatternFill("solid",fgColor=clr) if txt else rf,align=LW)
        ws.row_dimensions[ri].height=60
    widths=[16,45,14,8,12]+[35]*max_comments
    for ci,w in enumerate(widths,start=1): ws.column_dimensions[get_column_letter(ci)].width=w
    ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions; wb.save(filepath)


# ══════════════════════════════════════════════════════════════════════════════
#  DATE PICKER  (mini calendar widget)
# ══════════════════════════════════════════════════════════════════════════════
class DatePicker(tk.Toplevel):
    """Compact pop-up calendar. Returns selected date via .result (datetime)."""

    DAY_NAMES = ["Mo","Tu","We","Th","Fr","Sa","Su"]

    def __init__(self, parent, title="Pick a date", initial: datetime | None = None):
        super().__init__(parent)
        self.title(title)
        self.resizable(False, False)
        self.configure(bg=C["sky"])
        self.grab_set()          # modal
        self.result: datetime | None = None

        today = datetime.today()
        init  = initial or today
        self._year  = tk.IntVar(value=init.year)
        self._month = tk.IntVar(value=init.month)
        self._sel_day: int | None = init.day

        self._build()
        self._center(parent)

    def _center(self, parent):
        self.update_idletasks()
        x = parent.winfo_rootx() + (parent.winfo_width()  - self.winfo_width())  // 2
        y = parent.winfo_rooty() + (parent.winfo_height() - self.winfo_height()) // 2
        self.geometry(f"+{x}+{y}")

    def _build(self):
        # ── Nav bar ──────────────────────────────────────────────────────────
        nav = tk.Frame(self, bg=C["purple"])
        nav.pack(fill="x")

        tk.Button(nav, text="◀", command=self._prev_month,
                  bg=C["purple"], fg=C["white"], relief="flat",
                  font=("Comic Sans MS",11,"bold"), cursor="hand2", bd=0,
                  activebackground=C["pink"]).pack(side="left", padx=6, pady=6)

        self._title_lbl = tk.Label(nav, text="", bg=C["purple"], fg=C["white"],
                                   font=("Comic Sans MS",12,"bold"), width=16)
        self._title_lbl.pack(side="left", expand=True)

        tk.Button(nav, text="▶", command=self._next_month,
                  bg=C["purple"], fg=C["white"], relief="flat",
                  font=("Comic Sans MS",11,"bold"), cursor="hand2", bd=0,
                  activebackground=C["pink"]).pack(side="right", padx=6, pady=6)

        # ── Day-name header ───────────────────────────────────────────────────
        dh = tk.Frame(self, bg=C["lavender"])
        dh.pack(fill="x")
        for d in self.DAY_NAMES:
            tk.Label(dh, text=d, width=4, font=("Comic Sans MS",9,"bold"),
                     bg=C["lavender"], fg=C["purple"]).pack(side="left", padx=1, pady=2)

        # ── Grid ─────────────────────────────────────────────────────────────
        self._grid_frame = tk.Frame(self, bg=C["sky"])
        self._grid_frame.pack(padx=6, pady=6)

        # ── OK / Cancel ───────────────────────────────────────────────────────
        btn_row = tk.Frame(self, bg=C["sky"])
        btn_row.pack(fill="x", padx=6, pady=(0,8))
        tk.Button(btn_row, text="✅  OK", command=self._ok,
                  bg=C["green"], fg=C["white"], relief="flat",
                  font=("Comic Sans MS",10,"bold"), padx=12, pady=4,
                  cursor="hand2", activebackground=C["teal"]).pack(side="left", padx=4)
        tk.Button(btn_row, text="❌  Cancel", command=self.destroy,
                  bg=C["red"], fg=C["white"], relief="flat",
                  font=("Comic Sans MS",10,"bold"), padx=12, pady=4,
                  cursor="hand2", activebackground=C["orange"]).pack(side="left", padx=4)

        self._refresh()

    def _refresh(self):
        import calendar
        y, m = self._year.get(), self._month.get()
        self._title_lbl.config(text=datetime(y, m, 1).strftime("%B %Y"))

        # clear grid
        for w in self._grid_frame.winfo_children():
            w.destroy()

        cal = calendar.monthcalendar(y, m)
        today = datetime.today().date()

        for week in cal:
            row_f = tk.Frame(self._grid_frame, bg=C["sky"])
            row_f.pack()
            for day in week:
                if day == 0:
                    tk.Label(row_f, text="", width=4, bg=C["sky"]).pack(side="left", padx=1, pady=1)
                else:
                    is_sel   = (day == self._sel_day)
                    is_today = (datetime(y, m, day).date() == today)
                    bg   = C["purple"] if is_sel else (C["blue"] if is_today else C["white"])
                    fg   = C["white"]  if (is_sel or is_today) else C["text"]
                    font = ("Comic Sans MS",10,"bold") if is_sel else ("Comic Sans MS",9)
                    btn = tk.Button(row_f, text=str(day), width=3,
                                    bg=bg, fg=fg, relief="flat", font=font,
                                    activebackground=C["pink"], activeforeground=C["white"],
                                    cursor="hand2",
                                    command=lambda d=day: self._pick(d))
                    btn.pack(side="left", padx=1, pady=1)

    def _pick(self, day):
        self._sel_day = day
        self._refresh()

    def _prev_month(self):
        y, m = self._year.get(), self._month.get()
        m -= 1
        if m < 1: m, y = 12, y-1
        self._month.set(m); self._year.set(y)
        self._sel_day = None; self._refresh()

    def _next_month(self):
        y, m = self._year.get(), self._month.get()
        m += 1
        if m > 12: m, y = 1, y+1
        self._month.set(m); self._year.set(y)
        self._sel_day = None; self._refresh()

    def _ok(self):
        if self._sel_day is None:
            messagebox.showwarning("Pick a day!", "Please click on a day first! 📅", parent=self)
            return
        self.result = datetime(self._year.get(), self._month.get(), self._sel_day)
        self.destroy()


# ══════════════════════════════════════════════════════════════════════════════
#  FLOATING SPARKLES CANVAS
# ══════════════════════════════════════════════════════════════════════════════
class SparkleCanvas(tk.Canvas):
    SYMBOLS=["★","✦","✧","◆","✿","❋","♦","●"]
    def __init__(self,parent,**kw):
        super().__init__(parent,highlightthickness=0,**kw)
        self._stars=[]
        self.after(200,self._init)
    def _init(self):
        w,h=max(self.winfo_width(),800),max(self.winfo_height(),100)
        for _ in range(28):
            self._stars.append({
                "x":random.uniform(0,w),"y":random.uniform(0,h),
                "vy":random.uniform(-0.5,-0.15),"vx":random.uniform(-0.2,0.2),
                "size":random.randint(9,20),"sym":random.choice(self.SYMBOLS),
                "color":random.choice(RAINBOW),"phase":random.uniform(0,6.28),
                "dphase":random.uniform(0.03,0.09),
            })
        self._tick()
    def _tick(self):
        self.delete("sp")
        w,h=max(self.winfo_width(),800),max(self.winfo_height(),100)
        for s in self._stars:
            s["x"]=(s["x"]+s["vx"])%w; s["y"]+=s["vy"]
            if s["y"]<-24: s["y"]=h+10
            s["phase"]+=s["dphase"]
            sz=s["size"]+int(3*math.sin(s["phase"]))
            self.create_text(s["x"],s["y"],text=s["sym"],font=("Arial",sz),fill=s["color"],tags="sp")
        self.after(40,self._tick)


# ══════════════════════════════════════════════════════════════════════════════
#  PILL BUTTON
# ══════════════════════════════════════════════════════════════════════════════
class PillButton(tk.Canvas):
    def __init__(self,parent,text,command,color,**kw):
        super().__init__(parent,highlightthickness=0,cursor="hand2",bg=parent.cget("bg"),**kw)
        self._text=text; self._cmd=command; self._color=color
        self._phase=0.0; self._pressed=False
        self.bind("<Button-1>",self._press)
        self.bind("<ButtonRelease-1>",self._release)
        self.bind("<Configure>",lambda _:self._draw())
        self.after(20,self._animate)

    def _lgt(self,h,a=55):
        h=h.lstrip("#"); r,g,b=int(h[0:2],16),int(h[2:4],16),int(h[4:6],16)
        return f"#{min(255,r+a):02x}{min(255,g+a):02x}{min(255,b+a):02x}"
    def _drk(self,h,a=30):
        h=h.lstrip("#"); r,g,b=int(h[0:2],16),int(h[2:4],16),int(h[4:6],16)
        return f"#{max(0,r-a):02x}{max(0,g-a):02x}{max(0,b-a):02x}"

    def _pill(self,x1,y1,x2,y2,r,**kw):
        self.create_arc(x1,y1,x1+2*r,y1+2*r,start=90,extent=90,style="pieslice",**kw)
        self.create_arc(x2-2*r,y1,x2,y1+2*r,start=0,extent=90,style="pieslice",**kw)
        self.create_arc(x1,y2-2*r,x1+2*r,y2,start=180,extent=90,style="pieslice",**kw)
        self.create_arc(x2-2*r,y2-2*r,x2,y2,start=270,extent=90,style="pieslice",**kw)
        self.create_rectangle(x1+r,y1,x2-r,y2,**kw)
        self.create_rectangle(x1,y1+r,x2,y2-r,**kw)

    def _draw(self):
        self.delete("all")
        w,h=self.winfo_width(),self.winfo_height()
        if w<10 or h<10: return
        g=int(5+4*math.sin(self._phase)); r=h//2
        fill=self._drk(self._color) if self._pressed else self._color
        self._pill(g,g,w-2,h-2,r,fill="#c0bce8",outline="")
        self._pill(0,0,w-g-2,h-g-2,r,fill=fill,outline="")
        self._pill(8,5,w-g-10,h//2-2,r-4,fill=self._lgt(fill,70),outline="")
        self.create_text((w-g)//2,(h-g)//2,text=self._text,
                         font=("Comic Sans MS",14,"bold"),fill="#ffffff")

    def _animate(self):
        self._phase+=0.06; self._draw(); self.after(30,self._animate)

    def _press(self,_): self._pressed=True; self._cmd(); self._draw()
    def _release(self,_): self._pressed=False; self._draw()

    def update_label(self,t): self._text=t; self._draw()
    def update_color(self,c): self._color=c; self._draw()


# ══════════════════════════════════════════════════════════════════════════════
#  TOOLTIP
# ══════════════════════════════════════════════════════════════════════════════
def attach_tip(widget,text):
    tip=[None]
    def show(_):
        if tip[0]: return
        w=tk.Toplevel(widget); w.wm_overrideredirect(True)
        x=widget.winfo_rootx()+22; y=widget.winfo_rooty()+widget.winfo_height()+4
        w.wm_geometry(f"+{x}+{y}")
        tk.Label(w,text=text,bg="#fffde7",fg=C["text"],font=("Arial",9),
                 relief="solid",bd=1,padx=10,pady=6,wraplength=300).pack()
        tip[0]=w
    def hide(_):
        if tip[0]: tip[0].destroy(); tip[0]=None
    widget.bind("<Enter>",show); widget.bind("<Leave>",hide)


# ══════════════════════════════════════════════════════════════════════════════
#  CARD HELPER
# ══════════════════════════════════════════════════════════════════════════════
def make_card(parent,bg,title_emoji,title,subtitle,pady=10):
    outer=tk.Frame(parent,bg=bg,relief="flat")
    outer.pack(fill="x",pady=(0,pady))
    inner=tk.Frame(outer,bg=bg)
    inner.pack(fill="x",padx=14,pady=12)
    hdr=tk.Frame(inner,bg=bg); hdr.pack(fill="x",pady=(0,4))
    tk.Label(hdr,text=title_emoji,font=("Arial",18),bg=bg).pack(side="left",padx=(0,6))
    tk.Label(hdr,text=title,font=("Comic Sans MS",12,"bold"),bg=bg,fg=C["text"]).pack(side="left")
    if subtitle:
        tk.Label(hdr,text=f"  — {subtitle}",font=("Arial",8),bg=bg,fg=C["muted"]).pack(side="left")
    return inner


# ══════════════════════════════════════════════════════════════════════════════
#  DATE RANGE ROW  (reusable widget)
# ══════════════════════════════════════════════════════════════════════════════
class DateRangeRow(tk.Frame):
    """Displays  📅 From [DD MMM YYYY] ▸ 📅 To [DD MMM YYYY]  with calendar buttons."""

    FMT = "%d %b %Y"   # e.g. 01 Jan 2025

    def __init__(self, parent, bg, **kw):
        super().__init__(parent, bg=bg, **kw)
        today = datetime.today()
        six_months_ago = today - timedelta(days=180)

        self._from: datetime = six_months_ago
        self._to:   datetime = today

        self._from_var = tk.StringVar(value=self._from.strftime(self.FMT))
        self._to_var   = tk.StringVar(value=self._to.strftime(self.FMT))

        # ── FROM ──────────────────────────────────────────────────────────────
        tk.Label(self, text="📅 From:", font=("Comic Sans MS",10,"bold"),
                 bg=bg, fg=C["text"]).pack(side="left")

        self._from_btn = tk.Button(
            self, textvariable=self._from_var,
            command=self._pick_from, cursor="hand2",
            bg=C["blue"], fg=C["white"], relief="flat",
            font=("Comic Sans MS",10,"bold"), padx=10, pady=4,
            activebackground=C["purple"], activeforeground=C["white"],
        )
        self._from_btn.pack(side="left", padx=(4,0))
        attach_tip(self._from_btn, "Click to open a calendar and pick the START date.\nOnly posts from this date onwards will be collected.")

        tk.Label(self, text="  ➜  ", font=("Arial",14),
                 bg=bg, fg=C["muted"]).pack(side="left")

        # ── TO ────────────────────────────────────────────────────────────────
        tk.Label(self, text="📅 To:", font=("Comic Sans MS",10,"bold"),
                 bg=bg, fg=C["text"]).pack(side="left")

        self._to_btn = tk.Button(
            self, textvariable=self._to_var,
            command=self._pick_to, cursor="hand2",
            bg=C["pink"], fg=C["white"], relief="flat",
            font=("Comic Sans MS",10,"bold"), padx=10, pady=4,
            activebackground=C["purple"], activeforeground=C["white"],
        )
        self._to_btn.pack(side="left", padx=(4,0))
        attach_tip(self._to_btn, "Click to open a calendar and pick the END date.\nOnly posts up to this date will be collected.")

        # ── Quick preset buttons ───────────────────────────────────────────────
        presets_frame = tk.Frame(self, bg=bg)
        presets_frame.pack(side="left", padx=(14,0))
        tk.Label(presets_frame, text="Quick:", font=("Arial",8),
                 bg=bg, fg=C["muted"]).pack(side="left", padx=(0,4))
        for label, days in [("7d",7),("30d",30),("90d",90),("1yr",365),("All",0)]:
            b=tk.Button(presets_frame, text=label,
                       command=lambda d=days: self._preset(d),
                       bg=C["shadow"], fg=C["text"], relief="flat",
                       font=("Arial",8,"bold"), padx=5, pady=2,
                       cursor="hand2", activebackground=C["yellow"])
            b.pack(side="left", padx=2)
            tip_txt = f"Last {label}" if days else "All time (no date filter)"
            attach_tip(b, tip_txt)

    def _pick_from(self):
        dp = DatePicker(self, "📅 Pick START date", initial=self._from)
        self.wait_window(dp)
        if dp.result:
            if dp.result > self._to:
                messagebox.showwarning("Oops! 😅",
                    "Start date can't be after the end date!\nPick an earlier date. 📅", parent=self)
                return
            self._from = dp.result
            self._from_var.set(self._from.strftime(self.FMT))

    def _pick_to(self):
        dp = DatePicker(self, "📅 Pick END date", initial=self._to)
        self.wait_window(dp)
        if dp.result:
            if dp.result < self._from:
                messagebox.showwarning("Oops! 😅",
                    "End date can't be before the start date!\nPick a later date. 📅", parent=self)
                return
            self._to = dp.result
            self._to_var.set(self._to.strftime(self.FMT))

    def _preset(self, days):
        today = datetime.today().replace(hour=23, minute=59, second=59)
        if days == 0:
            self._from = datetime(2005, 1, 1)   # Reddit founding year
            self._to   = today
        else:
            self._from = today - timedelta(days=days)
            self._to   = today
        self._from_var.set(self._from.strftime(self.FMT))
        self._to_var.set(self._to.strftime(self.FMT))

    @property
    def date_from(self) -> datetime:
        return self._from.replace(hour=0, minute=0, second=0, microsecond=0)

    @property
    def date_to(self) -> datetime:
        return self._to.replace(hour=23, minute=59, second=59, microsecond=0)


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN APP
# ══════════════════════════════════════════════════════════════════════════════
class FunScraperApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("🎉 GenAI Ads Reddit Scraper — FUN EDITION!")
        self.geometry("1090x920"); self.minsize(920,760)
        self.configure(bg=C["sky"])
        self.update_idletasks()
        sw,sh=self.winfo_screenwidth(),self.winfo_screenheight()
        self.geometry(f"1090x920+{(sw-1090)//2}+{(sh-920)//2}")
        self._log_q    = queue.Queue()
        self._prog_q   = queue.Queue()
        self._running  = False
        self._stop_evt = threading.Event()
        self._build(); self._poll()

    # ── BUILD ─────────────────────────────────────────────────────────────────
    def _build(self):
        self._hdr(); self._body(); self._logpane()

    def _hdr(self):
        hdr=tk.Frame(self,bg=C["sky"],height=96); hdr.pack(fill="x"); hdr.pack_propagate(False)
        sc=SparkleCanvas(hdr,bg=C["sky"]); sc.place(x=0,y=0,relwidth=1,relheight=1)
        tf=tk.Frame(hdr,bg=C["sky"]); tf.place(relx=0.5,rely=0.5,anchor="center")
        row=tk.Frame(tf,bg=C["sky"]); row.pack()
        for txt,col in [("🌟","#ffd93d"),(" GenAI ",C["purple"]),("Ads ",C["pink"]),
                        ("Reddit ",C["blue"]),("Scraper ","#ff8c42"),("🚀",C["green"])]:
            tk.Label(row,text=txt,font=("Comic Sans MS",26,"bold"),bg=C["sky"],fg=col).pack(side="left")
        tk.Label(tf,text="✨  Fill in the form below, then click  GENERATE REPORT  to collect Reddit posts! ✨",
                 font=("Comic Sans MS",10),bg=C["sky"],fg=C["muted"]).pack()

    def _body(self):
        outer=tk.Frame(self,bg=C["sky"]); outer.pack(fill="both",expand=True,padx=16,pady=(6,0))
        cv=tk.Canvas(outer,bg=C["sky"],highlightthickness=0)
        vsb=ttk.Scrollbar(outer,orient="vertical",command=cv.yview)
        cv.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right",fill="y"); cv.pack(side="left",fill="both",expand=True)
        sf=tk.Frame(cv,bg=C["sky"]); fid=cv.create_window((0,0),window=sf,anchor="nw")
        sf.bind("<Configure>",lambda _:cv.configure(scrollregion=cv.bbox("all")))
        cv.bind("<Configure>",lambda e:cv.itemconfig(fid,width=e.width))
        cv.bind_all("<MouseWheel>",lambda e:cv.yview_scroll(int(-1*(e.delta/120)),"units"))
        self._form(sf)

    def _form(self, parent):
        cols=tk.Frame(parent,bg=C["sky"]); cols.pack(fill="x",pady=(4,0))
        left=tk.Frame(cols,bg=C["sky"]); right=tk.Frame(cols,bg=C["sky"])
        left.pack(side="left",fill="both",expand=True,padx=(0,8))
        right.pack(side="left",fill="both",expand=True)

        # ── LEFT: Subreddits ──────────────────────────────────────────────────
        si=make_card(left,C["lavender"],"📋","Subreddits","comma-separated, no r/ needed")
        self._note(si,"Type subreddit names separated by commas.  Example:  advertising, FacebookAds, GoogleAds",C["lavender"])
        self.sub_text=self._textbox(si,", ".join(DEFAULT_SUBREDDITS),C["lavender"],4)

        # ── LEFT: Keywords ────────────────────────────────────────────────────
        ki=make_card(left,C["mint"],"🔍","Search Keywords","one keyword or phrase per line")
        self._note(ki,"Each keyword is searched in every subreddit above. More keywords = more posts found.",C["mint"])
        self.kw_text=self._textbox(ki,"\n".join(DEFAULT_KEYWORDS),C["mint"],7)

        # ── LEFT: Date Range  ─────────────────────────────────────────────────
        di=make_card(left,C["lemon"],"📅","Date Range","only collect posts between these two dates")
        self._note(di,
            "Click the blue/pink date buttons to open a calendar 🗓️  "
            "and pick your start & end dates.  "
            "Or use the Quick buttons (7d / 30d / 90d / 1yr / All).",
            C["lemon"])
        self.date_row = DateRangeRow(di, bg=C["lemon"])
        self.date_row.pack(fill="x", pady=(4,0))

        # ── RIGHT: Parameters ─────────────────────────────────────────────────
        pi=make_card(right,C["peach"],"⚙️","Scrape Settings","control how much data to collect")
        g=tk.Frame(pi,bg=C["peach"]); g.pack(fill="x")
        g.columnconfigure(1,weight=1)

        self.max_posts_var    = tk.IntVar(value=50)
        self.max_comments_var = tk.IntVar(value=10)
        self.sort_var         = tk.StringVar(value="relevance")
        self.time_var         = tk.StringVar(value="all")

        spinfields=[
            ("📦 Max Posts / Subreddit", self.max_posts_var, 1, 500,
             "How many posts per subreddit?\n50 is a great start! More posts = more time."),
            ("💬 Comments per Post",     self.max_comments_var, 1, 25,
             "Each comment gets its OWN column in Excel.\n10 = 10 comment columns."),
        ]
        for rn,(lbl,var,lo,hi,tip) in enumerate(spinfields):
            l=tk.Label(g,text=lbl,font=("Comic Sans MS",10,"bold"),bg=C["peach"],fg=C["text"],anchor="w")
            l.grid(row=rn*2,column=0,sticky="w",pady=(10,0)); attach_tip(l,tip)
            sb=tk.Spinbox(g,textvariable=var,from_=lo,to=hi,width=8,
                          bg=C["white"],fg=C["text"],relief="flat",
                          font=("Comic Sans MS",11,"bold"),
                          buttonbackground=C["orange"],
                          highlightthickness=2,highlightcolor=C["orange"],
                          insertbackground=C["text"])
            sb.grid(row=rn*2,column=1,sticky="w",pady=(10,0),padx=(14,0))
            tk.Label(g,text=tip.split("\n")[0],font=("Arial",8),
                     bg=C["peach"],fg=C["muted"],anchor="w").grid(row=rn*2+1,column=0,columnspan=2,sticky="w",padx=4)

        dropfields=[
            ("🔀 Sort Results By",   self.sort_var, SORT_OPTIONS,
             "relevance = best keyword match\nnew = newest posts first\ntop = highest upvotes\ncomments = most discussed"),
            ("🕐 Reddit Time Filter", self.time_var, TIME_OPTIONS,
             "Reddit's own server-side time filter.\nUse 'all' — your Date Range above handles filtering."),
        ]
        for rn,(lbl,var,opts,tip) in enumerate(dropfields,start=len(spinfields)):
            l=tk.Label(g,text=lbl,font=("Comic Sans MS",10,"bold"),bg=C["peach"],fg=C["text"],anchor="w")
            l.grid(row=rn*2,column=0,sticky="w",pady=(10,0)); attach_tip(l,tip)
            dd=ttk.Combobox(g,textvariable=var,values=opts,state="readonly",
                            width=14,font=("Comic Sans MS",10))
            dd.grid(row=rn*2,column=1,sticky="w",pady=(10,0),padx=(14,0))
            tk.Label(g,text=tip.split("\n")[0],font=("Arial",8),
                     bg=C["peach"],fg=C["muted"],anchor="w").grid(row=rn*2+1,column=0,columnspan=2,sticky="w",padx=4)

        # ── RIGHT: Output ─────────────────────────────────────────────────────
        oi=make_card(right,C["lavender"],"💾","Output File","where should the Excel be saved?")
        self._note(oi,"📁 Folder to save in  (click Browse or type a path)",C["lavender"])
        pr=tk.Frame(oi,bg=C["lavender"]); pr.pack(fill="x",pady=(4,0))
        self.out_dir_var=tk.StringVar(value=str(Path.cwd()/"data"))
        pe=tk.Entry(pr,textvariable=self.out_dir_var,bg=C["white"],fg=C["text"],
                    relief="flat",font=("Comic Sans MS",10),bd=0,
                    highlightthickness=2,highlightcolor=C["purple"],insertbackground=C["text"])
        pe.pack(side="left",fill="x",expand=True,ipady=6,ipadx=6)
        attach_tip(pe,"The folder where the Excel file will be saved.\nIt will be created if it doesn't exist yet.")
        tk.Button(pr,text="📂 Browse…",command=self._browse,
                  bg=C["orange"],fg=C["white"],relief="flat",
                  font=("Comic Sans MS",10,"bold"),
                  activebackground=C["red"],activeforeground=C["white"],
                  padx=8,pady=4,cursor="hand2",bd=0).pack(side="left",padx=(6,0))
        self._note(oi,"📄 File name  (must end with .xlsx)",C["lavender"])
        self.fname_var=tk.StringVar(value="genai_ads_reddit.xlsx")
        fe=tk.Entry(oi,textvariable=self.fname_var,bg=C["white"],fg=C["text"],
                    relief="flat",font=("Comic Sans MS",10),bd=0,
                    highlightthickness=2,highlightcolor=C["purple"],insertbackground=C["text"])
        fe.pack(fill="x",ipady=6,ipadx=6,pady=(4,0))
        attach_tip(fe,"Name for your Excel report file.\nMake sure it ends with .xlsx")

        # ── ACTION BUTTONS — Generate + Stop side by side ─────────────────────
        btn_area = tk.Frame(right, bg=C["sky"])
        btn_area.pack(fill="x", pady=(10,0))

        # Generate (takes 70% width)
        gen_wrap = tk.Frame(btn_area, bg=C["sky"])
        gen_wrap.pack(side="left", fill="both", expand=True, padx=(0,6))
        self.run_btn = PillButton(gen_wrap, "🚀  GENERATE REPORT!", self._start, C["purple"], height=60)
        self.run_btn.pack(fill="x")

        # Stop (takes 30% width)
        stop_wrap = tk.Frame(btn_area, bg=C["sky"])
        stop_wrap.pack(side="left", fill="both", padx=(0,0))
        self.stop_btn = PillButton(stop_wrap, "🛑  STOP", self._stop, C["red"], height=60, width=130)
        self.stop_btn.pack(fill="x")

        # ── Progress bar ──────────────────────────────────────────────────────
        pf=tk.Frame(right,bg=C["sky"]); pf.pack(fill="x",pady=(8,0))
        self.prog_var=tk.DoubleVar(value=0)
        sty=ttk.Style(); sty.theme_use("clam")
        sty.configure("Candy.Horizontal.TProgressbar",
                      troughcolor=C["shadow"],background=C["pink"],
                      lightcolor=C["pink"],darkcolor=C["pink"],
                      bordercolor=C["sky"],thickness=16)
        ttk.Progressbar(pf,variable=self.prog_var,maximum=1.0,
                        style="Candy.Horizontal.TProgressbar").pack(fill="x")
        self.status_lbl=tk.Label(pf,
            text="👋 Ready! Fill in the form and hit the big purple button!",
            bg=C["sky"],fg=C["muted"],
            font=("Comic Sans MS",9),wraplength=500,justify="left")
        self.status_lbl.pack(anchor="w",pady=(4,0))

    def _logpane(self):
        tk.Frame(self,bg=C["shadow"],height=2).pack(fill="x",padx=16,pady=(4,0))
        lh=tk.Frame(self,bg=C["sky"]); lh.pack(fill="x",padx=16,pady=(4,0))
        tk.Label(lh,text="📜 Live Activity Log",font=("Comic Sans MS",10,"bold"),
                 bg=C["sky"],fg=C["purple"]).pack(side="left")
        tk.Button(lh,text="🧹 Clear",command=self._clrlog,
                  bg=C["shadow"],fg=C["text"],relief="flat",
                  font=("Comic Sans MS",8),padx=6,cursor="hand2").pack(side="right")
        self.log_box=ScrolledText(self,bg="#fafbff",fg=C["text"],
                                  font=("Consolas",9),relief="flat",bd=0,
                                  state="disabled",height=8)
        self.log_box.pack(fill="both",padx=16,pady=(4,12))
        self.log_box.tag_config("ok",   foreground=C["teal"])
        self.log_box.tag_config("warn", foreground=C["orange"])
        self.log_box.tag_config("head", foreground=C["purple"])
        self.log_box.tag_config("star", foreground=C["pink"])
        self.log_box.tag_config("info", foreground=C["text"])

    # ── HELPERS ───────────────────────────────────────────────────────────────
    def _note(self,p,t,bg):
        tk.Label(p,text=t,font=("Arial",8),bg=bg,fg=C["muted"],
                 anchor="w",wraplength=460,justify="left").pack(anchor="w",pady=(0,4))

    def _textbox(self,p,content,bg,height):
        t=tk.Text(p,bg=C["white"],fg=C["text"],insertbackground=C["text"],
                  relief="flat",font=("Consolas",10),height=height,
                  padx=8,pady=6,wrap="word",
                  highlightthickness=2,highlightcolor=C["purple"],
                  selectbackground=C["purple"],selectforeground=C["white"])
        t.insert("1.0",content); t.pack(fill="x",pady=(0,4)); return t

    def _browse(self):
        d=filedialog.askdirectory(title="📂 Pick your save folder!")
        if d: self.out_dir_var.set(d)

    def _clrlog(self):
        self.log_box.config(state="normal"); self.log_box.delete("1.0","end"); self.log_box.config(state="disabled")

    def _log(self,tag,msg):
        self.log_box.config(state="normal"); self.log_box.insert("end",msg+"\n",tag)
        self.log_box.see("end"); self.log_box.config(state="disabled")

    def _status(self,msg,col=None):
        self.status_lbl.config(text=msg,fg=col or C["muted"])

    # ── START ─────────────────────────────────────────────────────────────────
    def _start(self):
        if self._running: return

        subs=[s.strip().lstrip("r/").strip()
              for s in self.sub_text.get("1.0","end").replace("\n",",").split(",") if s.strip()]
        if not subs:
            messagebox.showerror("Oops! 😅","Please enter at least one subreddit! 📋"); return
        kws=[k.strip() for k in self.kw_text.get("1.0","end").splitlines() if k.strip()]
        if not kws:
            messagebox.showerror("Oops! 😅","Please enter at least one keyword! 🔍"); return

        date_from = self.date_row.date_from
        date_to   = self.date_row.date_to
        if date_from > date_to:
            messagebox.showerror("Oops! 😅",
                "Start date is after end date!\nPlease fix the date range. 📅"); return

        fname=self.fname_var.get().strip()
        if not fname.endswith(".xlsx"): fname+=".xlsx"
        out_dir=Path(self.out_dir_var.get().strip())
        out_path=str(out_dir/fname)

        config={
            "subreddits":   subs,
            "keywords":     kws,
            "max_posts":    self.max_posts_var.get(),
            "max_comments": self.max_comments_var.get(),
            "date_from":    date_from,
            "date_to":      date_to,
            "sort_by":      self.sort_var.get(),
            "time_filter":  self.time_var.get(),
        }

        self._stop_evt.clear()
        self._running = True
        self.run_btn.update_label("⏳  Working… please wait!")
        self.run_btn.update_color(C["orange"])
        self.prog_var.set(0)
        self._status(
            f"🚀 Scraping from {date_from.strftime('%d %b %Y')} "
            f"to {date_to.strftime('%d %b %Y')}… hang tight!",
            C["purple"]
        )
        self._log("head", f"🗓️  Date range: {date_from.strftime('%d %b %Y')} → {date_to.strftime('%d %b %Y')}")
        threading.Thread(target=self._worker, args=(config,out_path,out_dir), daemon=True).start()

    # ── STOP ──────────────────────────────────────────────────────────────────
    def _stop(self):
        if not self._running:
            self._status("😴 Nothing is running right now!", C["muted"])
            return
        if messagebox.askyesno("🛑 Stop Scraping?",
                "Are you sure you want to STOP?\n\n"
                "Any posts collected so far will still be saved to Excel! 💾"):
            self._stop_evt.set()
            self._status("🛑 Stop signal sent — finishing current post…", C["orange"])
            self._log("warn", "🛑 User pressed STOP — wrapping up…")

    # ── WORKER ────────────────────────────────────────────────────────────────
    def _worker(self,config,out_path,out_dir):
        try:
            out_dir.mkdir(parents=True,exist_ok=True)
            records=scrape(config, self._log_q, self._prog_q, self._stop_evt)
            if records:
                self._log_q.put(("info",f"\n💾 Saving to {out_path}…"))
                export_excel(records,out_path,config["max_comments"])
                self._log_q.put(("star",f"🎉 ALL DONE! {len(records)} posts saved to Excel!"))
                stopped = self._stop_evt.is_set()
                self._prog_q.put(("done", out_path, len(records), stopped))
            else:
                self._log_q.put(("warn","😕 No posts found — try different keywords or a wider date range."))
                self._prog_q.put(("empty",))
        except Exception as e:
            self._log_q.put(("warn",f"💥 Error: {e}"))
            self._prog_q.put(("error",str(e)))

    # ── POLL ──────────────────────────────────────────────────────────────────
    def _poll(self):
        while not self._log_q.empty():
            tag,msg=self._log_q.get_nowait(); self._log(tag,msg)
        while not self._prog_q.empty():
            item=self._prog_q.get_nowait()
            if isinstance(item,float):
                self.prog_var.set(item)
                self._status(f"⚡ Progress: {int(item*100)}%  Keep going!",C["purple"])
            elif isinstance(item,tuple):
                code=item[0]
                self.run_btn.update_label("🚀  GENERATE REPORT!")
                self.run_btn.update_color(C["purple"])
                self._running=False
                if code=="done":
                    _,path,n,stopped=item
                    self.prog_var.set(1.0)
                    msg = (f"⚠️ Stopped early — {n} posts saved → {path}" if stopped
                           else f"🎉 {n} posts saved → {path}")
                    self._status(msg, C["orange"] if stopped else C["green"])
                    popup_title = "⚠️ Stopped Early!" if stopped else "🎉 Yahoo!"
                    popup_msg   = (f"⚠️ Stopped early!\n\n✅ {n} posts still saved!\n\n📄 {path}\n\n👀 Open the output folder?"
                                   if stopped else
                                   f"✅ {n} posts saved!\n\n📄 {path}\n\n👀 Open the output folder?")
                    if messagebox.askyesno(popup_title, popup_msg):
                        import subprocess,platform
                        folder=str(Path(path).parent)
                        if platform.system()=="Windows": os.startfile(folder)
                        elif platform.system()=="Darwin": subprocess.Popen(["open",folder])
                        else: subprocess.Popen(["xdg-open",folder])
                elif code=="empty":
                    self._status("😕 No data found. Try broader keywords or wider date range!",C["orange"])
                else:
                    self._status(f"💥 Error: {item[1]}",C["red"])
        self.after(200,self._poll)


# ══════════════════════════════════════════════════════════════════════════════
if __name__=="__main__":
    app=FunScraperApp()
    app.mainloop()
